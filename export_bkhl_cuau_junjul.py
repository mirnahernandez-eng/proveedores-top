"""
export_bkhl_cuau_junjul.py
BKHL Cuautitlan (N1 + N2) — Junio 2026 y Julio hasta el 17.
3 hojas: CUAUTITLAN (N1+N2), NAVE 1, NAVE 2
Todos los 75 vendors del listado — vendors sin registros quedan en blanco.
"""
import re
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.cloud import bigquery

import sys

BASE     = Path(__file__).parent
BQ_TABLE = "wmt-intl-supplychain-gcp-prod.MR101_WM_AD_HOC.SCH_YMS_SEMANAL"

# Uso: python script.py FECHA_INI FECHA_FIN [nombre_salida.xlsx]
# Ej:  python script.py 2026-06-01 2026-06-30 bkhl_cuautitlan_junio.xlsx
if len(sys.argv) >= 3:
    FECHA_INI = sys.argv[1]
    FECHA_FIN = sys.argv[2]
    out_name  = sys.argv[3] if len(sys.argv) >= 4 else f"bkhl_cuautitlan_{FECHA_INI}_{FECHA_FIN}.xlsx"
else:
    FECHA_INI = "2026-06-01"
    FECHA_FIN = "2026-07-17"
    out_name  = "bkhl_cuautitlan_jun_jul17.xlsx"

OUT_FILE = BASE / out_name

CUAU_N1  = [7494]
CUAU_N2  = [7464]
CUAU_ALL = CUAU_N1 + CUAU_N2

# ─── Vendor rules (display_name, [python_re_patterns], bq_keyword) ─────────────
VENDOR_RULES = [
    ("4E GLOBAL SAPI DE CV",                               ["4E GLOBAL"],               "4E GLOBAL"),
    ("ABSORMEX CMPC TISSUE S.A. DE CV",                    ["ABSORMEX", "CMPC TISSUE"], "ABSORMEX"),
    ("ACH FOODS MEXICO S DE RL DE CV",                     ["ACH FOODS"],               "ACH FOODS"),
    ("AJEMEX SA DE CV",                                    ["AJEMEX"],                  "AJEMEX"),
    ("AKSI HERRAMIENTAS SA DE CV",                         ["AKSI"],                    "AKSI"),
    ("BIO PAPPEL SCRIBE SA DE CV",                         ["SCRIBE", "BIO PAPPEL"],    "SCRIBE"),
    ("BONAFONT SA DE CV",                                  ["BONAFONT"],                "BONAFONT"),
    ("BRONCOLIN SA DE CV",                                 ["BRONCOLIN"],               "BRONCOLIN"),
    ("CESARFER SA DE CV",                                  ["CESARFER"],                "CESARFER"),
    ("CH & ML ELECTRIC MEXICO S DE RL DE CV",              ["CH.{0,4}ML.{0,4}ELEC"],   "CH.*ML.*ELEC"),
    ("CHURCH & DWIGHT S DE RL DE CV",                      ["CHURCH.{0,8}DWIGHT"],      "CHURCH.*DWIGHT"),
    ("CIA COMERCIAL HERDEZ SA DE CV",                      ["HERDEZ"],                  "HERDEZ"),
    ("CIA INTERNAC COMERCIO SAPI CV",                      ["INTERNAC.{0,12}COMERC"],   "INTERNAC.*COMERC"),
    ("COLCHONES WENDY SA DE CV",                           ["WENDY"],                   "WENDY"),
    ("COMERC PEPSICO MEXICO S RL CV",                      ["PEPSICO"],                 "PEPSICO"),
    ("COMERCIAL 100 MEXICAN SA DE CV",                     ["100 MEXICAN"],             "100 MEXICAN"),
    ("COMERCIALIZADORA ELORO SA",                          ["ELORO"],                   "ELORO"),
    ("COMPANIA COMERCIALIZADORA PRODIN CENTRO SA DE CV",   ["PRODIN"],                  "PRODIN"),
    ("CONAGRA FOODS MEXICO SA DE CV",                      ["CONAGRA"],                 "CONAGRA"),
    ("CONSERVAS LA COSTENA SA DE CV",                      ["COSTE.A", "COSTENA"],      "COSTENA"),
    ("CORPORACION GAIRET SA DE CV",                        ["GAIRET"],                  "GAIRET"),
    ("CRISA LIBBEY COMERCIAL S RL CV",                     ["CRISA", "LIBBEY"],         "CRISA"),
    ("CUETARA DISTRIBUCION SA DE CV",                      ["CUETARA"],                 "CUETARA"),
    ("DASAVENA GOURMET SA DE CV",                          ["DASAVENA"],                "DASAVENA"),
    ("EFFEM MEXICO INC Y CIA S NC CV",                     ["EFFEM"],                   "EFFEM"),
    ("EMBOTELLADORA NIAGARA SA DE CV",                     ["NIAGARA"],                 "NIAGARA"),
    ("ESPEJOS INTELIGENTES SA DE CV",                      ["ESPEJOS"],                 "ESPEJOS"),
    ("FAB DE JABON LA CORONA SA DE CV",                    ["LA CORONA"],               "LA CORONA"),
    ("FACTOR PESCA SA DE CV",                              ["FACTOR PESCA"],            "FACTOR PESCA"),
    ("FANTASY RUZ S.A. DE C.V.",                           ["FANTASY RUZ"],             "FANTASY RUZ"),
    ("FRABEL SA DE CV",                                    ["FRABEL"],                  "FRABEL"),
    ("GANAD PROD DE LECHE PURA SA DE CV",                  ["LECHE PURA"],              "LECHE PURA"),
    ("GRUPO TAIFELDS SA DE CV",                            ["TAIFELDS"],                "TAIFELDS"),
    ("HALEON CONSUMER S DE RL DE CV",                      ["HALEON"],                  "HALEON"),
    ("HAPPY FLOWER MEXICANA SA DE CV",                     ["HAPPY FLOWER"],            "HAPPY FLOWER"),
    ("HENKEL CAPITAL SA DE CV",                            ["HENKEL"],                  "HENKEL"),
    ("HFC PRESTIGE INTERNATIONAL S D",                     ["HFC PRESTIGE"],            "HFC PRESTIGE"),
    ("HISENSE MEXICO S DE RL DE CV",                       ["HISENSE"],                 "HISENSE"),
    ("IMPERCAUCHO SA DE CV",                               ["IMPERCAUCHO"],             "IMPERCAUCHO"),
    ("IMPULSORA CAMPIRANO SA DE CV",                       ["CAMPIRANO"],               "CAMPIRANO"),
    ("IND NAC DE DETERGENTES SA DE CV",                    ["NAC.{0,10}DETERG"],        "DETERGENTES"),
    ("INDUSTRIAS OVARB SA DE CV",                          ["OVARB"],                   "OVARB"),
    ("INDUSTRIAS SALCOM SA DE CV",                         ["SALCOM"],                  "SALCOM"),
    ("JESSY INTERNACIONAL SA DE CV",                       ["JESSY"],                   "JESSY"),
    ("JUVASA SERVICIOS SA DE CV",                          ["JUVASA"],                  "JUVASA"),
    ("KIMBERLY CLARK DE MEX SA DE CV",                     ["KIMBERLY"],                "KIMBERLY"),
    ("KSMV CAPITAL SAPI DE",                               ["KSMV"],                    "KSMV"),
    ("LA MASCOTA SA DE CV",                                ["LA MASCOTA"],              "LA MASCOTA"),
    ("MARCAS NESTLE SA DE CV",                             ["NESTLE"],                  "NESTLE"),
    ("MATTEL DE MEXICO SA DE CV",                          ["MATTEL"],                  "MATTEL"),
    ("MONDELEZ MEXICO S DE RL DE CV",                      ["MONDELEZ"],                "MONDELEZ"),
    ("NEWELL BRANDS DE MEXICO SA DE C",                    ["NEWELL"],                  "NEWELL"),
    ("PENAFIEL BEBIDAS SA DE CV",                          ["PE.AFIEL", "PENAFIEL"],    "PENAFIEL"),
    ("PLAYERAS SOURCE SA DE CV",                           ["PLAYERAS SOURCE"],         "PLAYERAS SOURCE"),
    ("POLYCHEM SA DE CV",                                  ["POLYCHEM"],                "POLYCHEM"),
    ("PROBEMEX SA DE CV",                                  ["PROBEMEX"],                "PROBEMEX"),
    ("PROCTER & GAMBLE MEXICO S DE RL DE CV",              ["PROCTER"],                 "PROCTER"),
    ("PROD ALIMENT LA MODERNA SA DE CV",                   ["LA MODERNA"],              "LA MODERNA"),
    ("PRODUCTOS INTERNACIONALES MABE SA DE CV",            ["MABE"],                    "MABE"),
    ("PROXIMO NATAL SAPI DE CV",                           ["PROXIMO"],                 "PROXIMO"),
    ("QUALAMEX SA DE CV",                                  ["QUALAMEX"],                "QUALAMEX"),
    ("RAMIREZ ZUNIGA LAURA",                               ["RAMIREZ ZUNIGA"],          "RAMIREZ ZUNIGA"),
    ("RECKITT BENCKISER MEXICO",                           ["RECKITT"],                 "RECKITT"),
    ("REGALOS SIGLO XXI SA DE CV",                         ["REGALOS SIGLO"],           "REGALOS SIGLO"),
    ("SANTA CLARA MERCANTIL DE PACHUCA S DE RL DE CV",     ["SANTA CLARA"],             "SANTA CLARA"),
    ("SANTUL HERRAMIENTAS SA DE CV",                       ["SANTUL"],                  "SANTUL"),
    ("SCHETTINO HNOS SRL DE CV",                           ["SCHETTINO"],               "SCHETTINO"),
    ("SERVICIOS NUTRICIONALES MEAD JOHNSON S DE RL DE CV", ["MEAD JOHNSON"],            "MEAD JOHNSON"),
    ("SUPER FOODS FACTORY",                                ["SUPER FOODS"],             "SUPER FOODS"),
    ("TECNISPICE SA DE CV",                                ["TECNISPICE"],              "TECNISPICE"),
    ("TRESMONTES LUCCHETTI MEX SA DE CV",                  ["TRESMONTES", "LUCCHETTI"],"TRESMONTES"),
    ("UNILEVER DE MEXICO S RL CV",                         ["UNILEVER"],                "UNILEVER"),
    ("VASCONIA BRANDS SA DE CV",                           ["VASCONIA"],                "VASCONIA"),
    ("VCT & DG MEXICO SA DE CV",                           ["VCT.{0,5}DG"],             "VCT"),
    ("VIDRIERA SANTOS SA DE CV",                           ["VIDRIERA SANTOS"],         "VIDRIERA SANTOS"),
]

VENDOR_NAMES = [r[0] for r in VENDOR_RULES]
_BQ_PAT      = "|".join("(?:" + r[2] + ")" for r in VENDOR_RULES)

NUM_COLS = [
    "LLEGADA_A_TRAFICO", "ABRIR_CORTINA", "CERRAR_CORTINA",
    "PAPER_W", "SALIDA_DE_CD", "DURACION_DE_SERVICIO",
    "RECIBO_HRS", "LOS_HRS",
]
COLS_DET = [
    "FECHA", "# CITA", "VENDOR", "CEDIS", "CEDIS_NOMBRE",
    "TIPO_CITA", "CITAS_CORRECTAS", "SW",
    "PUNTUALIDAD", "DIFERENCIA_MIN",
    "LLEGADA_A_TRAFICO", "ABRIR_CORTINA", "CERRAR_CORTINA",
    "PAPER_W", "SALIDA_DE_CD", "DURACION_DE_SERVICIO",
    "RECIBO_HRS", "LOS_HRS",
    # columnas de resumen puntualidad (solo en filas de total)
    "# ANTICIP", "% ANTICIP", "# A TIEMPO", "% A TIEMPO", "# DESPUES", "% DESPUES",
]
COL_W = {
    "FECHA": 13, "# CITA": 16, "VENDOR": 44, "CEDIS": 8,
    "CEDIS_NOMBRE": 20, "TIPO_CITA": 16, "CITAS_CORRECTAS": 10, "SW": 6,
    "PUNTUALIDAD": 18, "DIFERENCIA_MIN": 14,
    "LLEGADA_A_TRAFICO": 14, "ABRIR_CORTINA": 13, "CERRAR_CORTINA": 14,
    "PAPER_W": 10, "SALIDA_DE_CD": 13, "DURACION_DE_SERVICIO": 18,
    "RECIBO_HRS": 12, "LOS_HRS": 12,
    "# ANTICIP": 10, "% ANTICIP": 10, "# A TIEMPO": 10,
    "% A TIEMPO": 10, "# DESPUES": 10, "% DESPUES": 10,
}
LEFT_COLS  = {"FECHA", "VENDOR", "CEDIS_NOMBRE", "TIPO_CITA", "PUNTUALIDAD"}
PUNT_COLS  = {"# ANTICIP", "% ANTICIP", "# A TIEMPO", "% A TIEMPO", "# DESPUES", "% DESPUES"}

# Clasificacion de puntualidad
_ANTICIP = {"1 DIA ANTES", "12-24 HORAS ANTES", "6-12 HORAS ANTES", "1-6 HORAS ANTES"}
_ATIME   = {"A TIEMPO"}
# cualquier valor que contenga "DESPUES" o "TARDE" se considera tarde


def clasif_puntualidad(cita_vs: str | None) -> str:
    if not cita_vs or str(cita_vs).strip() == "":
        return "SIN DATO"
    v = str(cita_vs).upper().strip()
    if v in _ANTICIP:
        return "ANTICIPADA"
    if v in _ATIME:
        return "A TIEMPO"
    if "DESPU" in v or "TARDE" in v:
        return "DESPUES"
    return "SIN DATO"


# ─── BigQuery ──────────────────────────────────────────────────────────────────
def query_bkhl() -> pd.DataFrame:
    client  = bigquery.Client()
    cedis_s = ",".join(str(c) for c in CUAU_ALL)
    q = (
        "SELECT\n"
        "    ARRIVAL_DATE                                   AS FECHA,\n"
        "    APPOINTMENT_NBR                                AS `# CITA`,\n"
        "    VENDOR,\n"
        "    SAFE_CAST(CEDIS AS INT64)                      AS CEDIS,\n"
        "    NOMBRE_CEDIS                                   AS CEDIS_NOMBRE,\n"
        "    TIPO_CITA,\n"
        "    CITAS_CORRECTAS,\n"
        "    SW,\n"
        "    COALESCE(CITA_VS_LLEGADA, 'SIN DATO')        AS CITA_VS_LLEGADA,\n"
        "    ROUND(SAFE_CAST(DIFERENCIA AS FLOAT64), 2)   AS DIFERENCIA_MIN,\n"
        "    ROUND(SAFE_CAST(LLEGADA_A_TRAFICO    AS FLOAT64)/60,4) AS LLEGADA_A_TRAFICO,\n"
        "    ROUND(SAFE_CAST(ABRIR_CORTINA        AS FLOAT64)/60,4) AS ABRIR_CORTINA,\n"
        "    ROUND(SAFE_CAST(CERRAR_CORTINA       AS FLOAT64)/60,4) AS CERRAR_CORTINA,\n"
        "    ROUND(SAFE_CAST(PAPER_W              AS FLOAT64)/60,4) AS PAPER_W,\n"
        "    ROUND(SAFE_CAST(SALIDA_DE_CD         AS FLOAT64)/60,4) AS SALIDA_DE_CD,\n"
        "    ROUND(SAFE_CAST(DURACION_DE_SERVICIO AS FLOAT64)/60,4) AS DURACION_DE_SERVICIO,\n"
        "    ROUND((\n"
        "        COALESCE(SAFE_CAST(ABRIR_CORTINA  AS FLOAT64),0)+\n"
        "        COALESCE(SAFE_CAST(CERRAR_CORTINA AS FLOAT64),0)+\n"
        "        COALESCE(SAFE_CAST(PAPER_W        AS FLOAT64),0)\n"
        "    )/60,4) AS RECIBO_HRS,\n"
        "    ROUND((\n"
        "        COALESCE(SAFE_CAST(LLEGADA_A_TRAFICO    AS FLOAT64),0)+\n"
        "        COALESCE(SAFE_CAST(DURACION_DE_SERVICIO AS FLOAT64),0)+\n"
        "        COALESCE(SAFE_CAST(SALIDA_DE_CD         AS FLOAT64),0)\n"
        "    )/60,4) AS LOS_HRS\n"
        "FROM `" + BQ_TABLE + "`\n"
        "WHERE ARRIVAL_DATE BETWEEN '" + FECHA_INI + "' AND '" + FECHA_FIN + "'\n"
        "  AND UPPER(TRIM(TIPO_CITA)) IN ('BACKHAUL','CNV BACKHAUL','REPRO BKH')\n"
        "  AND SAFE_CAST(CEDIS AS INT64) IN (" + cedis_s + ")\n"
        "  AND REGEXP_CONTAINS(UPPER(VENDOR), r'" + _BQ_PAT + "')\n"
        "  AND (\n"
        "        COALESCE(SAFE_CAST(ABRIR_CORTINA        AS FLOAT64),0)+\n"
        "        COALESCE(SAFE_CAST(CERRAR_CORTINA       AS FLOAT64),0)+\n"
        "        COALESCE(SAFE_CAST(PAPER_W              AS FLOAT64),0)+\n"
        "        COALESCE(SAFE_CAST(LLEGADA_A_TRAFICO    AS FLOAT64),0)+\n"
        "        COALESCE(SAFE_CAST(DURACION_DE_SERVICIO AS FLOAT64),0)+\n"
        "        COALESCE(SAFE_CAST(SALIDA_DE_CD         AS FLOAT64),0)\n"
        "      ) > 0\n"
        "ORDER BY ARRIVAL_DATE, CEDIS, VENDOR\n"
    )
    return client.query(q).to_dataframe()


# ─── Matching & enrich ────────────────────────────────────────────────────────
def match_vendor(raw: str) -> str | None:
    raw_up = str(raw).upper()
    for display, patterns, _ in VENDOR_RULES:
        for pat in patterns:
            if re.search(pat, raw_up):
                return display
    return None


def enrich(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["VENDOR"]      = df["VENDOR"].apply(match_vendor)
    df = df[df["VENDOR"].notna()].copy()
    df["PUNTUALIDAD"] = df["CITA_VS_LLEGADA"].apply(clasif_puntualidad)
    df["VENDOR"] = pd.Categorical(df["VENDOR"], categories=VENDOR_NAMES, ordered=True)
    return df.sort_values(["VENDOR", "FECHA", "CEDIS"]).reset_index(drop=True)


def _puntualidad_counts(sub: pd.DataFrame, total: int) -> dict:
    """Cuenta y % de ANTICIPADA / A TIEMPO / DESPUES para un subconjunto."""
    base = total if total > 0 else 1
    n_a = int((sub["PUNTUALIDAD"] == "ANTICIPADA").sum())
    n_t = int((sub["PUNTUALIDAD"] == "A TIEMPO").sum())
    n_d = int((sub["PUNTUALIDAD"] == "DESPUES").sum())
    return {
        "# ANTICIP":  n_a,
        "% ANTICIP":  f"{n_a/base:.0%}",
        "# A TIEMPO": n_t,
        "% A TIEMPO": f"{n_t/base:.0%}",
        "# DESPUES":  n_d,
        "% DESPUES":  f"{n_d/base:.0%}",
    }


# ─── Totales (incluye vendors vacios como fila en blanco) ─────────────────────
def make_totals(df: pd.DataFrame) -> pd.DataFrame:
    rows, grand_avgs = [], {c: [] for c in NUM_COLS}

    for vendor in VENDOR_NAMES:
        sub = df[df["VENDOR"] == vendor]
        n   = len(sub)
        row = {
            "FECHA": f"TOTAL - {vendor}", "# CITA": n if n else "",
            "VENDOR": vendor, "CEDIS": "", "CEDIS_NOMBRE": "",
            "TIPO_CITA": "", "CITAS_CORRECTAS": "", "SW": "",
            "PUNTUALIDAD": "", "DIFERENCIA_MIN": "",
        }
        for col in NUM_COLS:
            row[col] = None

        if not sub.empty:
            row.update(_puntualidad_counts(sub, n))
            base = sub[sub["RECIBO_HRS"].notna() & (sub["RECIBO_HRS"] > 0)]
            if not base.empty:
                for col in ("ABRIR_CORTINA", "CERRAR_CORTINA", "PAPER_W", "DURACION_DE_SERVICIO"):
                    avg = round(base[col].fillna(0).mean(), 4)
                    row[col] = avg
                    grand_avgs[col].append(avg)
                row["RECIBO_HRS"] = round(
                    row["ABRIR_CORTINA"] + row["CERRAR_CORTINA"] + row["PAPER_W"], 4)
                grand_avgs["RECIBO_HRS"].append(row["RECIBO_HRS"])
                for col in ("LLEGADA_A_TRAFICO", "SALIDA_DE_CD", "LOS_HRS"):
                    vals = sub[col].dropna()
                    vals = vals[vals > 0]
                    avg  = round(vals.mean(), 4) if len(vals) else None
                    row[col] = avg
                    if avg is not None:
                        grand_avgs[col].append(avg)
        else:
            for pc in PUNT_COLS:
                row[pc] = ""

        rows.append(row)

    # TOTAL GENERAL
    grand = {
        "FECHA": "TOTAL GENERAL", "# CITA": len(df),
        "VENDOR": "", "CEDIS": "", "CEDIS_NOMBRE": "",
        "TIPO_CITA": "", "CITAS_CORRECTAS": "", "SW": "",
        "PUNTUALIDAD": "", "DIFERENCIA_MIN": "",
    }
    for col in NUM_COLS:
        avgs = grand_avgs[col]
        grand[col] = round(sum(avgs) / len(avgs), 4) if avgs else None
    if grand.get("ABRIR_CORTINA") is not None:
        grand["RECIBO_HRS"] = round(
            grand["ABRIR_CORTINA"] + grand["CERRAR_CORTINA"] + grand["PAPER_W"], 4)
    grand.update(_puntualidad_counts(df, len(df)))
    rows.append(grand)

    return pd.DataFrame(rows, columns=COLS_DET)


# ─── Helpers ──────────────────────────────────────────────────────────────────
def fmt_hrs(h) -> str:
    if h is None or (isinstance(h, float) and (np.isnan(h) or h <= 0)):
        return ""
    hh = int(h)
    mm = round((h - hh) * 60)
    if mm == 60:
        hh += 1; mm = 0
    return f"{hh}:{mm:02d}"


def _styles(hdr_hex: str, alt_hex: str, tot_hex: str) -> dict:
    thin = Side(style="thin", color="CCCCCC")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    return {
        "hdr":   PatternFill("solid", fgColor=hdr_hex),
        "alt":   PatternFill("solid", fgColor=alt_hex),
        "tot":   PatternFill("solid", fgColor=tot_hex),
        "empty": PatternFill("solid", fgColor="F5F5F5"),
        "grand": PatternFill("solid", fgColor="2E4057"),
        "wf":    Font(bold=True, color="FFFFFF", size=10),
        "hf":    Font(bold=True, color="FFFFFF", size=9),
        "bf":    Font(bold=True, size=9),
        "nf":    Font(size=9),
        "gf":    Font(bold=True, color="FFFFFF", size=9),
        "ef":    Font(italic=True, color="AAAAAA", size=9),
        "ctr":   Alignment(horizontal="center", vertical="center"),
        "lft":   Alignment(horizontal="left",   vertical="center"),
        "brd":   brd,
    }


# ─── Write sheet ──────────────────────────────────────────────────────────────
def write_sheet(ws, df_data: pd.DataFrame, df_tot: pd.DataFrame,
                titulo: str, hdr_hex: str, alt_hex: str, tot_hex: str,
                n_records: int):
    ws.title = titulo[:31]
    s = _styles(hdr_hex, alt_hex, tot_hex)
    ncols = len(COLS_DET)

    # Titulo
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(1, 1,
        f"BKHL - {titulo} | {FECHA_INI} al {FECHA_FIN} | "
        f"{n_records:,} registros | tiempos en HORAS (hh:mm)")
    t.fill = s["hdr"]; t.font = s["wf"]; t.alignment = s["ctr"]

    # Cabecera
    for j, col in enumerate(COLS_DET, 1):
        c = ws.cell(2, j, col)
        c.fill = s["hdr"]; c.font = s["hf"]
        c.alignment = s["ctr"]; c.border = s["brd"]

    # Datos
    row_i = 3
    prev_vendor = None
    _p_green  = PatternFill("solid", fgColor="E2EFDA")  # anticipada
    _p_yellow = PatternFill("solid", fgColor="FFEB9C")  # a tiempo
    _p_red    = PatternFill("solid", fgColor="FFC7CE")  # despues
    _f_green  = Font(color="375623", size=9)
    _f_yellow = Font(color="9C5700", size=9)
    _f_red    = Font(color="9C0006", size=9)

    for _, row in df_data.iterrows():
        vendor = row["VENDOR"]
        fill = s["alt"] if vendor != prev_vendor else None
        prev_vendor = vendor
        for j, col in enumerate(COLS_DET, 1):
            val = row.get(col)
            if col in NUM_COLS:
                cell_val = fmt_hrs(val) or None
            elif col in PUNT_COLS:   # estas solo van en filas de total
                cell_val = None
            else:
                cell_val = None if (val is None or (isinstance(val, float) and np.isnan(val))) else val
            dc = ws.cell(row_i, j, cell_val)
            # color especial para PUNTUALIDAD
            if col == "PUNTUALIDAD" and cell_val:
                if cell_val == "ANTICIPADA":
                    dc.fill = _p_green;  dc.font = _f_green
                elif cell_val == "A TIEMPO":
                    dc.fill = _p_yellow; dc.font = _f_yellow
                elif cell_val == "DESPUES":
                    dc.fill = _p_red;    dc.font = _f_red
                else:
                    dc.font = s["nf"]
                    if fill: dc.fill = fill
            else:
                dc.font = s["bf"] if col in ("VENDOR", "RECIBO_HRS", "LOS_HRS") else s["nf"]
                dc.alignment = s["lft"] if col in LEFT_COLS else s["ctr"]
                if fill:
                    dc.fill = fill
            dc.alignment = s["lft"] if col in LEFT_COLS else s["ctr"]
            dc.border = s["brd"]
        row_i += 1

    # Separador
    ws.row_dimensions[row_i].height = 5
    row_i += 1

    # Totales (incluye vendors vacios)
    vendors_con_datos = set(df_data["VENDOR"].unique())
    for _, row in df_tot.iterrows():
        is_grand = str(row.get("FECHA", "")).startswith("TOTAL GENERAL")
        vendor   = str(row.get("VENDOR", ""))
        is_empty = (not is_grand) and (vendor not in vendors_con_datos)

        if is_grand:
            fill = s["grand"]; font = s["gf"]
        elif is_empty:
            fill = s["empty"]; font = s["ef"]
        else:
            fill = s["tot"]; font = s["bf"]

        for j, col in enumerate(COLS_DET, 1):
            val = row.get(col)
            if col in NUM_COLS:
                cell_val = fmt_hrs(val) or None
            else:
                cell_val = None if (val is None or (isinstance(val, float) and np.isnan(val))) else val
            dc = ws.cell(row_i, j, cell_val)
            dc.fill = fill; dc.font = font
            dc.alignment = s["lft"] if col in LEFT_COLS else s["ctr"]
            dc.border = s["brd"]
        row_i += 1

    # Anchos y opciones
    for j, col in enumerate(COLS_DET, 1):
        ws.column_dimensions[get_column_letter(j)].width = COL_W.get(col, 12)
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 26
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}2"


# ─── Main ──────────────────────────────────────────────────────────────────────
print(f"Consultando BigQuery BKHL Cuautitlan: {FECHA_INI} al {FECHA_FIN} ...")
df_raw = query_bkhl()
print(f"  {len(df_raw):,} registros raw")

df = enrich(df_raw)
print(f"  {len(df):,} registros con vendors objetivo")

df_n1 = df[df["CEDIS"].isin(CUAU_N1)].copy()
df_n2 = df[df["CEDIS"].isin(CUAU_N2)].copy()
print(f"  CUAU total: {len(df):,}  |  N1: {len(df_n1):,}  |  N2: {len(df_n2):,}")

found = df["VENDOR"].nunique()
print(f"  {found} de {len(VENDOR_NAMES)} vendors con registros ({len(VENDOR_NAMES)-found} en blanco)")

tot_all = make_totals(df)
tot_n1  = make_totals(df_n1)
tot_n2  = make_totals(df_n2)

print(f"\nEscribiendo {OUT_FILE} ...")
wb = Workbook()

write_sheet(wb.active,         df,    tot_all, "CUAUTITLAN (N1+N2)",
            "833C00", "FCE4D6", "F4B183", len(df))
write_sheet(wb.create_sheet(), df_n1, tot_n1,  "NAVE 1",
            "7B3F00", "FDE9D9", "F4B183", len(df_n1))
write_sheet(wb.create_sheet(), df_n2, tot_n2,  "NAVE 2",
            "4A1942", "F5E6F5", "D5A6E0", len(df_n2))

wb.save(OUT_FILE)
print(f"Listo: {OUT_FILE}")
print(f"  Hoja CUAUTITLAN : {len(df):,} citas")
print(f"  Hoja NAVE 1     : {len(df_n1):,} citas")
print(f"  Hoja NAVE 2     : {len(df_n2):,} citas")
