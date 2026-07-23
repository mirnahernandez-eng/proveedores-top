"""
export_tiempos_recibo.py
Exporta tiempos de recibo (Directo + BKHL) por vendor y CEDIS a Excel.
- Directo: vendor_cedis_mes_FINAL.csv (2026, TIPO_CITA Proveedor)
- BKHL:    BigQuery SCH_YMS_SEMANAL (2026, CNV Backhaul / Backhaul)
"""
import pandas as pd
import numpy as np
from pathlib import Path

BASE = Path(__file__).parent

# ─── 1. VENDORS de interés ──────────────────────────────────────────────────
# Nombre display → keyword para buscar en columna VENDOR del CSV
VENDORS = {
    "BONAFONT SA CV":              "BONAFONT",
    "COMERC PEPSICO MEXICO S RL CV":"PEPSICO",
    "EMBOTELLAD NIAGARA D MX S RLCV":"NIAGARA",
    "ENVASADORA LA SUPREMA SA DE CV":"SUPREMA",
    "FRABEL SA DE CV":             "FRABEL",
    "HERDEZ SA DE CV":             "HERDEZ",
    "JUGOS DEL VALLE SAPI DE CV":  "JUGOS DEL VALLE",
    "KIMBERLY CLARK MEXICO SA B CV":"KIMBERLY",
    "MARCAS NESTLE SA CV":         "NESTLE",
    "MONDELEZ MEXICO S DE RL DE CV":"MONDELEZ",
    "PROCTER AND GAMBLE MEXICO INC":"PROCTER",
    "SANTA CLARA MERC PACHU S RL CV":"SANTA CLARA",
    "UNILEVER DE MEXICO S RL CV":  "UNILEVER",
}
VENDOR_LABELS = list(VENDORS.keys())

# ─── 2. CEDIS → Ciudad ──────────────────────────────────────────────────────
CD_GROUPS = {
    "Cuautitlán N1":  [7494],
    "Cuautitlán N2":  [7464],
    "Megapark SMO":   [6388],
    "Santa Bárbara":  [7457, 7482],
    "Chihuahua":      [4640, 5780],
    "Culiacán":       [4971, 7455, 7487],
    "Mexicali":       [4924, 6140],
    "Monterrey":      [4995, 7461, 7490, 8806],
    "Chalco":         [7459, 7471, 7505],
    "Guadalajara":    [5907, 6238, 7460, 7493],
    "Mérida":         [4188, 7103, 7506],
    "Villahermosa":   [6550, 7453, 7468],
}
CD_LABELS = list(CD_GROUPS.keys())
CEDIS_TO_CD = {c: cd for cd, cs in CD_GROUPS.items() for c in cs}

def label_vendor(csv_vendor: str) -> str | None:
    """Devuelve el label del vendor o None si no hay match."""
    v = str(csv_vendor).upper()
    for label, kw in VENDORS.items():
        if kw.upper() in v:
            return label
    return None

def weighted_avg(grp: pd.DataFrame) -> float | None:
    """Promedio ponderado de TOTAL_HRS por TOTAL_CITAS."""
    total_citas = grp["TOTAL_CITAS"].sum()
    if total_citas == 0:
        return None
    return (grp["TOTAL_HRS"] * grp["TOTAL_CITAS"]).sum() / total_citas

def build_matrix(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    df_raw debe tener: VENDOR (str), CEDIS (int), TOTAL_HRS (float), TOTAL_CITAS (int)
    Devuelve DataFrame 13 vendors × 12 CDs con tiempo promedio en horas.
    """
    df_raw = df_raw.copy()
    df_raw["VENDOR_LABEL"] = df_raw["VENDOR"].apply(label_vendor)
    df_raw = df_raw[df_raw["VENDOR_LABEL"].notna()].copy()
    df_raw["CD"] = df_raw["CEDIS"].map(CEDIS_TO_CD)
    df_raw = df_raw[df_raw["CD"].notna()].copy()

    matrix = pd.DataFrame(index=VENDOR_LABELS, columns=CD_LABELS, dtype=float)
    for (vlbl, cd), grp in df_raw.groupby(["VENDOR_LABEL", "CD"]):
        matrix.loc[vlbl, cd] = weighted_avg(grp)
    return matrix

# ─── 3. Datos DIRECTO desde CSV mensual ─────────────────────────────────────
print("Cargando Directo desde vendor_cedis_mes_FINAL.csv...")
df_dir = pd.read_csv(BASE / "vendor_cedis_mes_FINAL.csv")
# Solo 2026
if "ANIO" in df_dir.columns:
    df_dir = df_dir[df_dir["ANIO"] == 2026]
# TOTAL_HRS ya está en horas
mat_directo = build_matrix(df_dir)
print(f"  Filas raw: {len(df_dir)} | celdas con dato: {mat_directo.notna().sum().sum()}/156")

# ─── 4. Datos BKHL desde BigQuery ───────────────────────────────────────────
mat_bkhl = pd.DataFrame(index=VENDOR_LABELS, columns=CD_LABELS, dtype=float)
bkhl_ok = False
BQ_TABLE = "wmt-intl-supplychain-gcp-prod.MR101_WM_AD_HOC.SCH_YMS_SEMANAL"

print("Conectando a BigQuery para datos BKHL...")
try:
    from google.cloud import bigquery
    client = bigquery.Client()
    query = f"""
    SELECT
        CEDIS,
        VENDOR,
        COUNT(*) AS TOTAL_CITAS,
        AVG(
            COALESCE(SAFE_CAST(LLEGADA_A_TRAFICO AS FLOAT64), 0) +
            COALESCE(SAFE_CAST(DURACION_DE_SERVICIO AS FLOAT64), 0) +
            COALESCE(SAFE_CAST(SALIDA_DE_CD AS FLOAT64), 0)
        ) / 60.0 AS TOTAL_HRS
    FROM `{BQ_TABLE}`
    WHERE ANIO = 2026
      AND UPPER(TRIM(TIPO_CITA)) IN ('BACKHAUL', 'CNV BACKHAUL', 'REPRO BKH')
      AND CITAS_CORRECTAS = 1
      AND (
            COALESCE(SAFE_CAST(LLEGADA_A_TRAFICO AS FLOAT64), 0) +
            COALESCE(SAFE_CAST(DURACION_DE_SERVICIO AS FLOAT64), 0) +
            COALESCE(SAFE_CAST(SALIDA_DE_CD AS FLOAT64), 0)
          ) > 0
    GROUP BY CEDIS, VENDOR
    """
    print("  Ejecutando query BQ...")
    df_bkhl = client.query(query).to_dataframe()
    df_bkhl["CEDIS"] = pd.to_numeric(df_bkhl["CEDIS"], errors="coerce")
    mat_bkhl = build_matrix(df_bkhl)
    bkhl_ok = True
    print(f"  BQ devolvió {len(df_bkhl)} filas | celdas con dato: {mat_bkhl.notna().sum().sum()}/156")
except Exception as e:
    print(f"   BigQuery no disponible: {e}")
    print("  La hoja BKHL irá vacía — actualiza después con conexión BQ activa.")

# ─── 5. Formato en horas:minutos ────────────────────────────────────────────
def fmt_hrs(h):
    if pd.isna(h) or h is None:
        return ""
    hh = int(h)
    mm = round((h - hh) * 60)
    if mm == 60:
        hh += 1; mm = 0
    return f"{hh}:{mm:02d}"

mat_dir_fmt  = mat_directo.map(fmt_hrs)
mat_bkhl_fmt = mat_bkhl.map(fmt_hrs)

# ─── 6. Escribir Excel ──────────────────────────────────────────────────────
out_path = BASE / "tiempos_recibo_dir_bkhl.xlsx"
print(f"\nEscribiendo {out_path} ...")

HEADER_FILL_DIR  = "1F4E79"   # azul oscuro
HEADER_FILL_BKHL = "833C00"   # café/naranja
CELL_FILL_ALT    = "D9E1F2"   # azul muy claro (filas alternas)
CELL_FILL_ALT_B  = "FCE4D6"   # naranja muy claro (BKHL)
FONT_WHITE = "FFFFFF"
FONT_DARK  = "000000"

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

wb = Workbook()

def write_sheet(ws, mat: pd.DataFrame, tipo: str, hdr_hex: str, alt_hex: str):
    ws.title = tipo
    hdr_fill  = PatternFill("solid", fgColor=hdr_hex)
    hdr_font  = Font(bold=True, color=FONT_WHITE, size=10)
    alt_fill  = PatternFill("solid", fgColor=alt_hex)
    bold_font = Font(bold=True, size=10)
    norm_font = Font(size=10)
    center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left      = Alignment(horizontal="left",   vertical="center")
    thin      = Side(style="thin", color="CCCCCC")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header row 1: título tipo
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(CD_LABELS)+1)
    c = ws.cell(1, 1, f"TIEMPO DE RECIBO — {tipo.upper()} | 2026 (promedio ponderado por citas, en hh:mm)")
    c.fill = hdr_fill
    c.font = hdr_font
    c.alignment = center

    # Header row 2: Proveedor + CDs
    ws.cell(2, 1, "Proveedor").fill = hdr_fill
    ws.cell(2, 1).font = hdr_font
    ws.cell(2, 1).alignment = center
    ws.cell(2, 1).border = border
    for j, cd in enumerate(CD_LABELS, start=2):
        c = ws.cell(2, j, cd)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = center
        c.border = border

    # Data rows
    for i, vendor in enumerate(VENDOR_LABELS, start=3):
        use_alt = (i % 2 == 0)
        row_fill = alt_fill if use_alt else None

        vc = ws.cell(i, 1, vendor)
        vc.font = bold_font
        vc.alignment = left
        vc.border = border
        if row_fill: vc.fill = row_fill

        for j, cd in enumerate(CD_LABELS, start=2):
            val = mat.loc[vendor, cd]
            dc = ws.cell(i, j, val if val != "" else None)
            dc.font = norm_font
            dc.alignment = center
            dc.border = border
            if row_fill: dc.fill = row_fill

    # Column widths
    ws.column_dimensions["A"].width = 38
    for j in range(2, len(CD_LABELS)+2):
        ws.column_dimensions[get_column_letter(j)].width = 14
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = "B3"

# Sheet 1: Directo
ws1 = wb.active
write_sheet(ws1, mat_dir_fmt,  "DIRECTO", HEADER_FILL_DIR,  CELL_FILL_ALT)

# Sheet 2: BKHL
ws2 = wb.create_sheet()
write_sheet(ws2, mat_bkhl_fmt, "BKHL",    HEADER_FILL_BKHL, CELL_FILL_ALT_B)

if not bkhl_ok:
    from openpyxl.styles import Font as OFont
    note = ws2.cell(15, 1, " Datos BKHL no disponibles — sin conexión BigQuery al momento de generar.")
    note.font = OFont(bold=True, color="C00000", size=11)

wb.save(out_path)
print(f" Excel guardado: {out_path}")
print(f"   Hoja DIRECTO: {mat_directo.notna().sum().sum()} celdas con dato")
print(f"   Hoja BKHL:    {'OK' if bkhl_ok else 'vacía (sin BQ)'}")
