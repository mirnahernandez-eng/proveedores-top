"""
export_entrega_directa_mty_mxl.py
Citas de Entrega Directa (TIPO_CITA = PROVEEDOR / CITA NUEVA, es decir,
NO Backhaul) en los CEDIS de Monterrey (MTY) y Mexicali (MXL).
Rango: 2026-01-01 -> 2026-08-07 (mismo corte que el tablero LOS Proveedores TOP).
Una pestana por CEDIS, una fila por cita.
"""
import warnings; warnings.filterwarnings("ignore")
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.cloud import bigquery

BASE     = Path(__file__).parent
BQ_TABLE = "wmt-intl-supplychain-gcp-prod.MR101_WM_AD_HOC.SCH_YMS_SEMANAL"
OUT_FILE = BASE / "entrega_directa_mty_mxl.xlsx"
FECHA_INI, FECHA_FIN = "2026-01-01", "2026-08-07"

q = f"""
SELECT
    ARRIVAL_DATE                                             AS FECHA,
    APPOINTMENT_NBR                                          AS CITA,
    VENDOR,
    SAFE_CAST(CEDIS AS INT64)                                AS CEDIS,
    NOMBRE_CEDIS,
    LOCACION,
    TIPO_CITA,
    CITAS_CORRECTAS,
    SW,
    COALESCE(CITA_VS_LLEGADA, 'SIN DATO')                   AS PUNTUALIDAD,
    ROUND(SAFE_CAST(LLEGADA_A_TRAFICO    AS FLOAT64)/60, 4) AS LLEGADA_HRS,
    ROUND(SAFE_CAST(ABRIR_CORTINA        AS FLOAT64)/60, 4) AS ABRIR_CORTINA,
    ROUND(SAFE_CAST(CERRAR_CORTINA       AS FLOAT64)/60, 4) AS CERRAR_CORTINA,
    ROUND(SAFE_CAST(PAPER_W              AS FLOAT64)/60, 4) AS PAPER_W,
    ROUND(SAFE_CAST(SALIDA_DE_CD         AS FLOAT64)/60, 4) AS SALIDA_HRS,
    ROUND(SAFE_CAST(DURACION_DE_SERVICIO AS FLOAT64)/60, 4) AS SERVICIO_HRS,
    ROUND((
        COALESCE(SAFE_CAST(LLEGADA_A_TRAFICO    AS FLOAT64),0)+
        COALESCE(SAFE_CAST(DURACION_DE_SERVICIO AS FLOAT64),0)+
        COALESCE(SAFE_CAST(SALIDA_DE_CD         AS FLOAT64),0)
    )/60, 4) AS LOS_HRS,
    CASE
        WHEN UPPER(LOCACION) LIKE 'MONTE%' THEN 'MTY'
        WHEN UPPER(LOCACION) LIKE 'MEXIC%' THEN 'MXL'
    END AS CEDIS_COD
FROM `{BQ_TABLE}`
WHERE ARRIVAL_DATE BETWEEN '{FECHA_INI}' AND '{FECHA_FIN}'
  AND (UPPER(LOCACION) LIKE 'MONTE%' OR UPPER(LOCACION) LIKE 'MEXIC%')
  AND TRIM(UPPER(TIPO_CITA)) IN ('PROVEEDOR', 'CITA NUEVA')
ORDER BY CEDIS_COD, FECHA, VENDOR
"""

print(f"Consultando Entrega Directa MTY + MXL ({FECHA_INI} -> {FECHA_FIN}) ...")
df = bigquery.Client().query(q).to_dataframe()
print(f"  {len(df):,} citas de entrega directa encontradas")
print(df.groupby("CEDIS_COD").size().to_string())

# ─── Helpers ──────────────────────────────────────────────────────────────────
HRS_COLS = ["LLEGADA_HRS", "ABRIR_CORTINA", "CERRAR_CORTINA",
            "PAPER_W", "SALIDA_HRS", "SERVICIO_HRS", "LOS_HRS"]
COLS     = ["FECHA", "CITA", "VENDOR", "CEDIS", "NOMBRE_CEDIS", "TIPO_CITA",
            "CITAS_CORRECTAS", "SW", "PUNTUALIDAD",
            "LLEGADA_HRS", "ABRIR_CORTINA", "CERRAR_CORTINA", "PAPER_W",
            "SALIDA_HRS", "SERVICIO_HRS", "LOS_HRS"]
COL_W    = {
    "FECHA": 13, "CITA": 16, "VENDOR": 32, "CEDIS": 8, "NOMBRE_CEDIS": 22,
    "TIPO_CITA": 14, "CITAS_CORRECTAS": 10, "SW": 6, "PUNTUALIDAD": 18,
    "LLEGADA_HRS": 13, "ABRIR_CORTINA": 13, "CERRAR_CORTINA": 14,
    "PAPER_W": 10, "SALIDA_HRS": 12, "SERVICIO_HRS": 13, "LOS_HRS": 12,
}
LEFT_COLS = {"FECHA", "VENDOR", "NOMBRE_CEDIS", "TIPO_CITA", "PUNTUALIDAD"}
_ANTICIP  = {"1 DIA ANTES", "12-24 HORAS ANTES", "6-12 HORAS ANTES", "1-6 HORAS ANTES"}
thin      = Side(style="thin", color="CCCCCC")
BRD       = Border(left=thin, right=thin, top=thin, bottom=thin)
HDR_HEX   = {"MTY": "1D4ED8", "MXL": "0F766E"}
ALT_FILL  = PatternFill("solid", fgColor="F1F5F9")


def fmt_hrs(h) -> str:
    if h is None or h == "":
        return ""
    try:
        h = float(h)
    except (TypeError, ValueError):
        return ""
    if np.isnan(h) or h <= 0:
        return ""
    hh = int(h); mm = round((h - hh) * 60)
    if mm == 60:
        hh += 1; mm = 0
    return f"{hh}:{mm:02d}"


# Objetos de estilo pre-creados UNA sola vez y reutilizados en todas las celdas
# (crear un Font/Alignment/Border nuevo por celda es lo que hace que openpyxl
# se vuelva lentisimo con datasets grandes -- aqui son ~76k citas).
_F_NORMAL     = Font(size=9)
_F_BOLD       = Font(bold=True, size=9)
_AL_LEFT      = Alignment(horizontal="left", vertical="center")
_AL_CENTER    = Alignment(horizontal="center", vertical="center")
_FILL_ANTICIP = PatternFill("solid", fgColor="E2EFDA")
_FILL_TIEMPO  = PatternFill("solid", fgColor="FFEB9C")
_FILL_TARDE   = PatternFill("solid", fgColor="FFC7CE")
_F_ANTICIP    = Font(color="375623", size=9)
_F_TIEMPO     = Font(color="9C5700", size=9)
_F_TARDE      = Font(color="9C0006", size=9)


def punt_style(v):
    v = str(v).upper().strip() if v else ""
    if v in _ANTICIP:
        return _FILL_ANTICIP, _F_ANTICIP
    if v == "A TIEMPO":
        return _FILL_TIEMPO, _F_TIEMPO
    if "DESPU" in v or "TARDE" in v:
        return _FILL_TARDE, _F_TARDE
    return None, _F_NORMAL


def write_sheet(wb, cedis_cod, cedis_label, sub):
    ws = wb.create_sheet(cedis_cod)
    ncols = len(COLS)
    hdr_fill = PatternFill("solid", fgColor=HDR_HEX[cedis_cod])
    tot_fill = PatternFill("solid", fgColor="2E4057")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(1, 1,
        f"ENTREGA DIRECTA — {cedis_label} ({cedis_cod}) | {FECHA_INI} a {FECHA_FIN} | "
        f"{len(sub):,} citas | proveedor / cita nueva (sin backhaul)")
    t.fill = hdr_fill; t.font = Font(bold=True, color="FFFFFF", size=10)
    t.alignment = Alignment(horizontal="center", vertical="center")

    for j, col in enumerate(COLS, 1):
        c = ws.cell(2, j, col)
        c.fill = hdr_fill; c.font = Font(bold=True, color="FFFFFF", size=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BRD

    for i, row in enumerate(sub.itertuples(index=False), start=0):
        ri = i + 3
        row_fill = ALT_FILL if i % 2 == 0 else None
        for j, col in enumerate(COLS, 1):
            val = getattr(row, col)
            cell_val = (fmt_hrs(val) or None) if col in HRS_COLS else (
                None if (val is None or (isinstance(val, float) and np.isnan(val))) else val)
            dc = ws.cell(ri, j, cell_val)
            if col == "PUNTUALIDAD":
                pf, pfont = punt_style(val)
                dc.font = pfont
                if pf:
                    dc.fill = pf
                elif row_fill:
                    dc.fill = row_fill
            else:
                dc.font = _F_BOLD if col in ("CITA", "LOS_HRS") else _F_NORMAL
                if row_fill:
                    dc.fill = row_fill
            dc.alignment = _AL_LEFT if col in LEFT_COLS else _AL_CENTER
            # Sin borde por celda de datos (76k+ filas) -- el costo de crear/
            # asignar un Border por celda es el cuello de botella real de
            # openpyxl. El encabezado si lleva borde para que se vea limpio.

    ws.row_dimensions[len(sub) + 3].height = 5
    ri_tot = len(sub) + 4
    ws.cell(ri_tot, 1, f"PROMEDIO ({len(sub)} citas)").fill = tot_fill
    ws.cell(ri_tot, 1).font = Font(bold=True, color="FFFFFF", size=9)
    ws.cell(ri_tot, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(ri_tot, 1).border = BRD
    for j, col in enumerate(COLS, 1):
        if j == 1:
            continue
        cell_val = None
        if col in HRS_COLS:
            vals = sub[col].dropna(); vals = vals[vals > 0]
            cell_val = fmt_hrs(vals.mean()) if len(vals) else None
        dc = ws.cell(ri_tot, j, cell_val)
        dc.fill = tot_fill; dc.font = Font(bold=True, color="FFFFFF", size=9)
        dc.alignment = Alignment(horizontal="center", vertical="center")
        dc.border = BRD

    for j, col in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(j)].width = COL_W.get(col, 12)
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}2"


wb = Workbook()
wb.remove(wb.active)  # quitamos la hoja default vacia
for cedis_cod, cedis_label in [("MTY", "Monterrey"), ("MXL", "Mexicali")]:
    sub = df[df["CEDIS_COD"] == cedis_cod].reset_index(drop=True)
    write_sheet(wb, cedis_cod, cedis_label, sub)

wb.save(OUT_FILE)
print(f"\nListo -> {OUT_FILE.name}")
