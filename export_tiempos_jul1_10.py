"""
export_tiempos_jul1_10.py
Tiempos de recibo Jul 1-10 2026 (Directo + BKHL) desde BigQuery.
"""
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent
BQ_TABLE = "wmt-intl-supplychain-gcp-prod.MR101_WM_AD_HOC.SCH_YMS_SEMANAL"
FECHA_INI = "2026-07-01"
FECHA_FIN = "2026-07-10"

# ─── VENDORS ────────────────────────────────────────────────────────────────
VENDORS = {
    "BONAFONT SA CV":               "BONAFONT",
    "COMERC PEPSICO MEXICO S RL CV": "PEPSICO",
    "EMBOTELLAD NIAGARA D MX S RLCV":"NIAGARA",
    "ENVASADORA LA SUPREMA SA DE CV":"SUPREMA",
    "FRABEL SA DE CV":              "FRABEL",
    "HERDEZ SA DE CV":              "HERDEZ",
    "JUGOS DEL VALLE SAPI DE CV":   "JUGOS DEL VALLE",
    "KIMBERLY CLARK MEXICO SA B CV": "KIMBERLY",
    "MARCAS NESTLE SA CV":          "NESTLE",
    "MONDELEZ MEXICO S DE RL DE CV":"MONDELEZ",
    "PROCTER AND GAMBLE MEXICO INC":"PROCTER",
    "SANTA CLARA MERC PACHU S RL CV":"SANTA CLARA",
    "UNILEVER DE MEXICO S RL CV":   "UNILEVER",
}
VENDOR_LABELS = list(VENDORS.keys())

# ─── CEDIS agrupados por ciudad ──────────────────────────────────────────────
CD_GROUPS = {
    "Cuautitlan N1":  [7494],
    "Cuautitlan N2":  [7464],
    "Megapark SMO":   [6388],
    "Santa Barbara":  [7457, 7482],
    "Chihuahua":      [4640, 5780],
    "Culiacan":       [4971, 7455, 7487],
    "Mexicali":       [4924, 6140],
    "Monterrey":      [4995, 7461, 7490, 8806],
    "Chalco":         [7459, 7471, 7505],
    "Guadalajara":    [5907, 6238, 7460, 7493],
    "Merida":         [4188, 7103, 7506],
    "Villahermosa":   [6550, 7453, 7468],
}
CD_LABELS = list(CD_GROUPS.keys())
CEDIS_TO_CD = {c: cd for cd, cs in CD_GROUPS.items() for c in cs}

def label_vendor(v: str):
    v = str(v).upper()
    for label, kw in VENDORS.items():
        if kw.upper() in v:
            return label
    return None

def build_matrix(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["VENDOR_LABEL"] = df["VENDOR"].apply(label_vendor)
    df = df[df["VENDOR_LABEL"].notna()].copy()
    df["CEDIS"] = pd.to_numeric(df["CEDIS"], errors="coerce")
    df["CD"] = df["CEDIS"].map(CEDIS_TO_CD)
    df = df[df["CD"].notna()].copy()

    mat = pd.DataFrame(index=VENDOR_LABELS, columns=CD_LABELS, dtype=float)
    for (vlbl, cd), grp in df.groupby(["VENDOR_LABEL", "CD"]):
        total_c = grp["TOTAL_CITAS"].sum()
        if total_c > 0:
            mat.loc[vlbl, cd] = (grp["TOTAL_HRS"] * grp["TOTAL_CITAS"]).sum() / total_c
    return mat

def fmt_hrs(h):
    if pd.isna(h) or h is None:
        return ""
    hh = int(h)
    mm = round((h - hh) * 60)
    if mm == 60:
        hh += 1; mm = 0
    return f"{hh}:{mm:02d}"

# ─── BigQuery query generica ─────────────────────────────────────────────────
def query_bq(tipo_filter: str) -> pd.DataFrame:
    from google.cloud import bigquery
    client = bigquery.Client()
    q = f"""
    SELECT
        CEDIS,
        VENDOR,
        COUNT(*) AS TOTAL_CITAS,
        AVG(
            COALESCE(SAFE_CAST(LLEGADA_A_TRAFICO    AS FLOAT64), 0) +
            COALESCE(SAFE_CAST(DURACION_DE_SERVICIO AS FLOAT64), 0) +
            COALESCE(SAFE_CAST(SALIDA_DE_CD         AS FLOAT64), 0)
        ) / 60.0 AS TOTAL_HRS
    FROM `{BQ_TABLE}`
    WHERE ARRIVAL_DATE BETWEEN '{FECHA_INI}' AND '{FECHA_FIN}'
      AND {tipo_filter}
      AND CITAS_CORRECTAS = 1
      AND (
            COALESCE(SAFE_CAST(LLEGADA_A_TRAFICO    AS FLOAT64), 0) +
            COALESCE(SAFE_CAST(DURACION_DE_SERVICIO AS FLOAT64), 0) +
            COALESCE(SAFE_CAST(SALIDA_DE_CD         AS FLOAT64), 0)
          ) > 0
    GROUP BY CEDIS, VENDOR
    """
    return client.query(q).to_dataframe()

print(f"Consultando BigQuery: {FECHA_INI} al {FECHA_FIN}")

# Directo
print("  Directo (Proveedor / Cita Nueva)...")
df_dir = query_bq("UPPER(TRIM(TIPO_CITA)) IN ('PROVEEDOR', 'CITA NUEVA')")
mat_dir = build_matrix(df_dir)
print(f"    {len(df_dir)} filas | {mat_dir.notna().sum().sum()}/156 celdas")

# BKHL
print("  BKHL (Backhaul)...")
df_bkhl = query_bq("UPPER(TRIM(TIPO_CITA)) IN ('BACKHAUL', 'CNV BACKHAUL', 'REPRO BKH')")
mat_bkhl = build_matrix(df_bkhl)
print(f"    {len(df_bkhl)} filas | {mat_bkhl.notna().sum().sum()}/156 celdas")

# ─── Excel ───────────────────────────────────────────────────────────────────
def write_sheet(ws, mat: pd.DataFrame, tipo: str, hdr_hex: str, alt_hex: str):
    hdr_fill = PatternFill("solid", fgColor=hdr_hex)
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    alt_fill = PatternFill("solid", fgColor=alt_hex)
    bold_f   = Font(bold=True, size=10)
    norm_f   = Font(size=10)
    center   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_a   = Alignment(horizontal="left",   vertical="center")
    thin     = Side(style="thin", color="CCCCCC")
    brd      = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.title = tipo
    ncols = len(CD_LABELS) + 1

    # Titulo
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, f"TIEMPO DE RECIBO — {tipo} | {FECHA_INI} al {FECHA_FIN} (hh:mm, prom ponderado)")
    c.fill = hdr_fill; c.font = hdr_font; c.alignment = center

    # Cabecera CDs
    ws.cell(2, 1, "Proveedor").fill = hdr_fill
    ws.cell(2, 1).font = hdr_font
    ws.cell(2, 1).alignment = center
    ws.cell(2, 1).border = brd
    for j, cd in enumerate(CD_LABELS, 2):
        cell = ws.cell(2, j, cd)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = center; cell.border = brd

    # Datos
    for i, vendor in enumerate(VENDOR_LABELS, 3):
        use_alt = (i % 2 == 0)
        rf = alt_fill if use_alt else None
        vc = ws.cell(i, 1, vendor)
        vc.font = bold_f; vc.alignment = left_a; vc.border = brd
        if rf: vc.fill = rf
        for j, cd in enumerate(CD_LABELS, 2):
            val = mat.loc[vendor, cd]
            dc = ws.cell(i, j, fmt_hrs(val) if not pd.isna(val) else None)
            dc.font = norm_f; dc.alignment = center; dc.border = brd
            if rf: dc.fill = rf

    ws.column_dimensions["A"].width = 38
    for j in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(j)].width = 14
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = "B3"

out = BASE / "tiempos_recibo_jul1_10.xlsx"
wb = Workbook()
write_sheet(wb.active,        mat_dir,  "DIRECTO", "1F4E79", "D9E1F2")
write_sheet(wb.create_sheet(), mat_bkhl, "BKHL",   "833C00", "FCE4D6")
wb.save(out)
print(f"\nExcel guardado: {out}")
