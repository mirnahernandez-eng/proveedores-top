"""
export_cuau_top75_bkhl_ed.py
Cuautitlan 4 hojas: CUAUTITLAN (N1+N2), NAVE 1, NAVE 2, NAVE 3 (7492)
BKHL + Entregas Directas — TOP 75 vendors en orden exacto.
Genera DOS archivos: junio 2026 y julio 1-10 2026.
"""
import re, warnings
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.cloud import bigquery

warnings.filterwarnings("ignore")

BASE     = Path(__file__).parent
BQ_TABLE = "wmt-intl-supplychain-gcp-prod.MR101_WM_AD_HOC.SCH_YMS_SEMANAL"

CEDIS_N1  = [7494]
CEDIS_N2  = [7464]
CEDIS_N3  = [7492]
CEDIS_ALL = CEDIS_N1 + CEDIS_N2 + CEDIS_N3

TIPOS_BKHL = "'BACKHAUL','CNV BACKHAUL','REPRO BKH'"
TIPOS_ED   = "'PROVEEDOR','CITA NUEVA'"

FECHA_INICIO_GLOBAL = "2026-06-01"
FECHA_FIN_GLOBAL    = "2026-07-10"

# ─── 75 Vendor rules ──────────────────────────────────────────────────────────
VENDOR_RULES = [
    ("BONAFONT SA CV",                  ["BONAFONT"],                "BONAFONT"),
    ("COMERC PEPSICO MEXICO S RL CV",   ["PEPSICO"],                 "PEPSICO"),
    ("EMBOTELLAD NIAGARA D MX S RLCV",  ["NIAGARA"],                 "NIAGARA"),
    ("ENVASADORA LA SUPREMA SA DE CV",  ["SUPREMA"],                 "SUPREMA"),
    ("FRABEL SA DE CV",                 ["FRABEL"],                  "FRABEL"),
    ("HERDEZ SA DE CV",                 ["HERDEZ"],                  "HERDEZ"),
    ("JUGOS DEL VALLE SAPI DE CV",      ["JUGOS.{0,8}VALLE","DEL VALLE SAPI"], "DEL VALLE"),
    ("KIMBERLY CLARK MEXICO SA B CV",   ["KIMBERLY"],                "KIMBERLY"),
    ("MARCAS NESTLE SA CV",             ["NESTLE"],                  "NESTLE"),
    ("MONDELEZ MEXICO S DE RL DE CV",   ["MONDELEZ"],                "MONDELEZ"),
    ("PROCTER AND GAMBLE MEXICO INC",   ["PROCTER"],                 "PROCTER"),
    ("SANTA CLARA MERC PACHU S RL CV",  ["SANTA CLARA"],             "SANTA CLARA"),
    ("UNILEVER DE MEXICO S RL CV",      ["UNILEVER"],                "UNILEVER"),
    ("4E GLOBAL SAPI DE CV",            ["4E GLOBAL"],               "4E GLOBAL"),
    ("ABSORMEX CMPC TISSUE SA DE CV",   ["ABSORMEX", "CMPC TISSUE"], "ABSORMEX"),
    ("ACH FOODS MEXICO S DE RL DE CV",  ["ACH FOODS"],               "ACH FOODS"),
    ("AJEMEX SA DE CV",                 ["AJEMEX"],                  "AJEMEX"),
    ("AKSI HERRAMIENTAS SA DE CV",      ["AKSI"],                    "AKSI"),
    ("BIO PAPPEL SCRIBE SA DE CV",      ["SCRIBE", "BIO PAPPEL"],    "SCRIBE"),
    ("BRONCOLIN SA DE CV",              ["BRONCOLIN"],               "BRONCOLIN"),
    ("CESARFER SA CV",                  ["CESARFER"],                "CESARFER"),
    ("CHURCH & DWIGHT SRL CV",          ["CHURCH.{0,8}DWIGHT"],      "CHURCH.*DWIGHT"),
    ("CIA COMERC PRODIN CENTRO SA CV",  ["PRODIN"],                  "PRODIN"),
    ("CIA INTERNAC COMERCIO SAPI CV",   ["INTERNAC.{0,12}COMERC"],   "INTERNAC.*COMERC"),
    ("COLCHONES WENDY SA DE CV",        ["WENDY"],                   "WENDY"),
    ("COMERCIALIZADORA EL ORO SA",      ["ELORO"],                   "ELORO"),
    ("CONAGRA FOODS MEXICO SA CV",      ["CONAGRA"],                 "CONAGRA"),
    ("CONSERVAS LA COSTENA SA CV",      ["COSTE.A", "COSTENA"],      "COSTENA"),
    ("CORPORACION GAIRET SA DE CV",     ["GAIRET"],                  "GAIRET"),
    ("CRISA LIBBEY MEX S DE RL DE CV",  ["CRISA", "LIBBEY"],         "CRISA"),
    ("CUETARA DISTRIBUCION SA DE CV",   ["CUETARA"],                 "CUETARA"),
    ("DASAVENA GOURMET SA DE CV",       ["DASAVENA"],                "DASAVENA"),
    ("EFFEM MEXICO INC Y CIA S NC CV",  ["EFFEM"],                   "EFFEM"),
    ("ESPEJOS INTELIGENTES SA DE CV",   ["ESPEJOS"],                 "ESPEJOS"),
    ("FAB DE JABON LA CORONA SA CV",    ["LA CORONA"],               "LA CORONA"),
    ("FACTOR PESCA SA DE CV",           ["FACTOR PESCA"],            "FACTOR PESCA"),
    ("FANTASY RUZ SA DE CV",            ["FANTASY RUZ"],             "FANTASY RUZ"),
    ("GANAD PROD DE LECHE PURA SA CV",  ["LECHE PURA"],              "LECHE PURA"),
    ("GRUPO TAIFELDS SA DE CV",         ["TAIFELDS"],                "TAIFELDS"),
    ("HAPPY FLOWER MEXICANA SA CV",     ["HAPPY FLOWER"],            "HAPPY FLOWER"),
    ("HENKEL CAPITAL SA DE CV",         ["HENKEL"],                  "HENKEL"),
    ("HFC PRESTIGE INTERNATIONAL S D",  ["HFC PRESTIGE"],            "HFC PRESTIGE"),
    ("HISENSE MEXICO S DE RL DE CV",    ["HISENSE"],                 "HISENSE"),
    ("IMPERCAUCHO SA DE CV",            ["IMPERCAUCHO"],             "IMPERCAUCHO"),
    ("IMPULSORA CAMPIRANO SA DE CV",    ["CAMPIRANO"],               "CAMPIRANO"),
    ("IND NAC DE DETERGENTES SA CV",    ["NAC.{0,10}DETERG"],        "DETERGENTES"),
    ("INDUSTRIAS OVARB SA CV",          ["OVARB"],                   "OVARB"),
    ("INDUSTRIAS SALCOM SA CV",         ["SALCOM"],                  "SALCOM"),
    ("JESSY INTERNACIONAL SA DE CV",    ["JESSY"],                   "JESSY"),
    ("KSMV CAPITAL SAPI DE CV",         ["KSMV"],                    "KSMV"),
    ("LA MASCOTA SA CV",                ["LA MASCOTA"],              "LA MASCOTA"),
    ("MATTEL DE MEXICO SA CV",          ["MATTEL"],                  "MATTEL"),
    ("MEAD JOHNSON NUTRIC MEX SRLCV",   ["MEAD JOHNSON"],            "MEAD JOHNSON"),
    ("NEWELL BRANDS DE MEXICO SA CV",   ["NEWELL"],                  "NEWELL"),
    ("PENAFIEL BEBIDAS SA DE CV",       ["PE.AFIEL", "PENAFIEL"],    "PENAFIEL"),
    ("POLYCHEM SA CV",                  ["POLYCHEM"],                "POLYCHEM"),
    ("PROBEMEX SA DE CV",               ["PROBEMEX"],                "PROBEMEX"),
    ("PROD ALIMENT LA MODERNA SA CV",   ["LA MODERNA"],              "LA MODERNA"),
    ("PROD INTERNACIONALES MABE SACV",  ["MABE"],                    "MABE"),
    ("PROXIMO NATAL SAPI DE CV",        ["PROXIMO"],                 "PROXIMO"),
    ("QUALAMEX SA CV",                  ["QUALAMEX"],                "QUALAMEX"),
    ("RAMIREZ ZUNIGA LAURA",            ["RAMIREZ ZUNIGA"],          "RAMIREZ ZUNIGA"),
    ("RECKITT BENCKISER MEXICO SA CV",  ["RECKITT"],                 "RECKITT"),
    ("REGALOS SIGLO XXI SA DE CV",      ["REGALOS SIGLO"],           "REGALOS SIGLO"),
    ("SANTUL HERRAMIENTAS SA DE CV",    ["SANTUL"],                  "SANTUL"),
    ("SCHETTINO HNOS SRL CV",           ["SCHETTINO"],               "SCHETTINO"),
    ("SUPER FOODS FACTORY SA DE CV",    ["SUPER FOODS"],             "SUPER FOODS"),
    ("TECNISPICE SA DE CV",             ["TECNISPICE"],              "TECNISPICE"),
    ("TRESMONTES LUCCHETTI MEX SA CV",  ["TRESMONTES", "LUCCHETTI"], "TRESMONTES"),
    ("VASCONIA BRANDS SA DE CV",        ["VASCONIA"],                "VASCONIA"),
    ("VCT & DG MEXICO SA DE CV",        ["VCT.{0,5}DG"],             "VCT"),
    ("VIDRIERA SANTOS SA DE CV",        ["VIDRIERA SANTOS"],         "VIDRIERA SANTOS"),
    ("CH & ML ELECTRIC MEXICO S DE R",  ["CH.{0,4}ML.{0,4}ELEC"],   "CH.*ML.*ELEC"),
    ("COMERCIALIZADORA 100 MEXICANA",   ["100 MEXICAN"],             "100 MEXICAN"),
    ("PLAYERAS SOURCE SA CV",           ["PLAYERAS SOURCE"],         "PLAYERAS SOURCE"),
]

VENDOR_NAMES = [r[0] for r in VENDOR_RULES]   # orden exacto del usuario
_BQ_PAT      = "|".join("(?:" + r[2] + ")" for r in VENDOR_RULES)

# ─── Columnas ─────────────────────────────────────────────────────────────────
NUM_BKHL  = ["LLEGADA_A_TRAFICO","ABRIR_CORTINA","CERRAR_CORTINA",
             "PAPER_W","SALIDA_DE_CD","DURACION_DE_SERVICIO","RECIBO_HRS","LOS_HRS"]
PUNT_COLS = {"# ANTICIP","% ANTICIP","# A TIEMPO","% A TIEMPO","# DESPUES","% DESPUES"}
TOT_ONLY  = PUNT_COLS | {"# CITAS_ED","RECIBO_ED_AVG"}

COLS_DET = [
    "FECHA","# CITA","VENDOR","CEDIS","CEDIS_NOMBRE",
    "TIPO_CITA","CITAS_CORRECTAS","SW",
    "PUNTUALIDAD","DIFERENCIA_MIN",
    "LLEGADA_A_TRAFICO","ABRIR_CORTINA","CERRAR_CORTINA",
    "PAPER_W","SALIDA_DE_CD","DURACION_DE_SERVICIO",
    "RECIBO_HRS","RECIBO_ED","LOS_HRS",
    "# ANTICIP","% ANTICIP","# A TIEMPO","% A TIEMPO","# DESPUES","% DESPUES",
    "# CITAS_ED","RECIBO_ED_AVG",
]
COL_W = {
    "FECHA":13,"# CITA":16,"VENDOR":44,"CEDIS":8,
    "CEDIS_NOMBRE":20,"TIPO_CITA":16,"CITAS_CORRECTAS":10,"SW":6,
    "PUNTUALIDAD":18,"DIFERENCIA_MIN":14,
    "LLEGADA_A_TRAFICO":14,"ABRIR_CORTINA":13,"CERRAR_CORTINA":14,
    "PAPER_W":10,"SALIDA_DE_CD":13,"DURACION_DE_SERVICIO":18,
    "RECIBO_HRS":13,"RECIBO_ED":14,"LOS_HRS":12,
    "# ANTICIP":10,"% ANTICIP":10,"# A TIEMPO":10,
    "% A TIEMPO":10,"# DESPUES":10,"% DESPUES":10,
    "# CITAS_ED":11,"RECIBO_ED_AVG":14,
}
LEFT_COLS = {"FECHA","VENDOR","CEDIS_NOMBRE","TIPO_CITA","PUNTUALIDAD"}

_ANTICIP = {"1 DIA ANTES","12-24 HORAS ANTES","6-12 HORAS ANTES","1-6 HORAS ANTES"}
_ATIME   = {"A TIEMPO"}


def clasif_punt(v) -> str:
    if not v or str(v).strip() == "":
        return "SIN DATO"
    v = str(v).upper().strip()
    if v in _ANTICIP:                    return "ANTICIPADA"
    if v in _ATIME:                      return "A TIEMPO"
    if "DESPU" in v or "TARDE" in v:     return "DESPUES"
    return "SIN DATO"


# ─── BigQuery ─────────────────────────────────────────────────────────────────
def query_bq(tipos_str: str) -> pd.DataFrame:
    cedis_s = ",".join(str(c) for c in CEDIS_ALL)
    q = (
        "SELECT ARRIVAL_DATE AS FECHA, APPOINTMENT_NBR AS `# CITA`, VENDOR,\n"
        "  SAFE_CAST(CEDIS AS INT64) AS CEDIS, NOMBRE_CEDIS AS CEDIS_NOMBRE,\n"
        "  TIPO_CITA, CITAS_CORRECTAS, SW,\n"
        "  COALESCE(CITA_VS_LLEGADA,'SIN DATO') AS CITA_VS_LLEGADA,\n"
        "  ROUND(SAFE_CAST(DIFERENCIA AS FLOAT64),2) AS DIFERENCIA_MIN,\n"
        "  ROUND(SAFE_CAST(LLEGADA_A_TRAFICO    AS FLOAT64)/60,4) AS LLEGADA_A_TRAFICO,\n"
        "  ROUND(SAFE_CAST(ABRIR_CORTINA        AS FLOAT64)/60,4) AS ABRIR_CORTINA,\n"
        "  ROUND(SAFE_CAST(CERRAR_CORTINA       AS FLOAT64)/60,4) AS CERRAR_CORTINA,\n"
        "  ROUND(SAFE_CAST(PAPER_W              AS FLOAT64)/60,4) AS PAPER_W,\n"
        "  ROUND(SAFE_CAST(SALIDA_DE_CD         AS FLOAT64)/60,4) AS SALIDA_DE_CD,\n"
        "  ROUND(SAFE_CAST(DURACION_DE_SERVICIO AS FLOAT64)/60,4) AS DURACION_DE_SERVICIO,\n"
        "  ROUND((COALESCE(SAFE_CAST(ABRIR_CORTINA  AS FLOAT64),0)+\n"
        "         COALESCE(SAFE_CAST(CERRAR_CORTINA AS FLOAT64),0)+\n"
        "         COALESCE(SAFE_CAST(PAPER_W        AS FLOAT64),0))/60,4) AS RECIBO_HRS,\n"
        "  ROUND((COALESCE(SAFE_CAST(LLEGADA_A_TRAFICO    AS FLOAT64),0)+\n"
        "         COALESCE(SAFE_CAST(DURACION_DE_SERVICIO AS FLOAT64),0)+\n"
        "         COALESCE(SAFE_CAST(SALIDA_DE_CD         AS FLOAT64),0))/60,4) AS LOS_HRS\n"
        f"FROM `{BQ_TABLE}`\n"
        f"WHERE ARRIVAL_DATE BETWEEN '{FECHA_INICIO_GLOBAL}' AND '{FECHA_FIN_GLOBAL}'\n"
        f"  AND UPPER(TRIM(TIPO_CITA)) IN ({tipos_str})\n"
        f"  AND SAFE_CAST(CEDIS AS INT64) IN ({cedis_s})\n"
        f"  AND REGEXP_CONTAINS(UPPER(VENDOR), r'{_BQ_PAT}')\n"
        "  AND (COALESCE(SAFE_CAST(ABRIR_CORTINA        AS FLOAT64),0)+\n"
        "       COALESCE(SAFE_CAST(CERRAR_CORTINA       AS FLOAT64),0)+\n"
        "       COALESCE(SAFE_CAST(PAPER_W              AS FLOAT64),0)+\n"
        "       COALESCE(SAFE_CAST(LLEGADA_A_TRAFICO    AS FLOAT64),0)+\n"
        "       COALESCE(SAFE_CAST(DURACION_DE_SERVICIO AS FLOAT64),0)+\n"
        "       COALESCE(SAFE_CAST(SALIDA_DE_CD         AS FLOAT64),0)) > 0\n"
        "ORDER BY ARRIVAL_DATE, CEDIS, VENDOR\n"
    )
    return bigquery.Client().query(q).to_dataframe()


def match_vendor(raw: str):
    raw_up = str(raw).upper()
    for display, patterns, _ in VENDOR_RULES:
        for pat in patterns:
            if re.search(pat, raw_up):
                return display
    return None


def enrich(df: pd.DataFrame, tipo_label: str) -> pd.DataFrame:
    df = df.copy()
    df["VENDOR"]      = df["VENDOR"].apply(match_vendor)
    df = df[df["VENDOR"].notna()].copy()
    df["PUNTUALIDAD"] = df["CITA_VS_LLEGADA"].apply(clasif_punt)
    df["_TIPO"]       = tipo_label
    df["RECIBO_ED"]   = np.nan
    return df.reset_index(drop=True)


def filter_sheet(df: pd.DataFrame, cedis: list) -> pd.DataFrame:
    out = df[df["CEDIS"].isin(cedis)].copy()
    out["VENDOR"] = pd.Categorical(out["VENDOR"], categories=VENDOR_NAMES, ordered=True)
    return out.sort_values(["VENDOR","FECHA","CEDIS"]).reset_index(drop=True)


# ─── Totales ──────────────────────────────────────────────────────────────────
def _punt_counts(sub: pd.DataFrame, total: int) -> dict:
    base = total if total > 0 else 1
    n_a  = int((sub["PUNTUALIDAD"] == "ANTICIPADA").sum())
    n_t  = int((sub["PUNTUALIDAD"] == "A TIEMPO").sum())
    n_d  = int((sub["PUNTUALIDAD"] == "DESPUES").sum())
    return {
        "# ANTICIP": n_a, "% ANTICIP": f"{n_a/base:.0%}",
        "# A TIEMPO":n_t, "% A TIEMPO":f"{n_t/base:.0%}",
        "# DESPUES": n_d, "% DESPUES": f"{n_d/base:.0%}",
    }


def make_totals(df_b: pd.DataFrame, df_e: pd.DataFrame) -> pd.DataFrame:
    rows, gavg = [], {c: [] for c in NUM_BKHL}
    g_ed_hrs   = []

    for vendor in VENDOR_NAMES:
        sb = df_b[df_b["VENDOR"] == vendor]
        se = df_e[df_e["VENDOR"] == vendor]
        n  = len(sb)
        row = {
            "FECHA": f"TOTAL - {vendor}", "# CITA": n if n else "",
            "VENDOR": vendor, "CEDIS": "", "CEDIS_NOMBRE": "",
            "TIPO_CITA": "BKHL", "CITAS_CORRECTAS": "", "SW": "",
            "PUNTUALIDAD": "", "DIFERENCIA_MIN": "", "_TIPO": "TOT",
            "RECIBO_ED": None,
        }
        for c in NUM_BKHL:
            row[c] = None

        if not sb.empty:
            row.update(_punt_counts(sb, n))
            base = sb[sb["RECIBO_HRS"].notna() & (sb["RECIBO_HRS"] > 0)]
            if not base.empty:
                for c in ("ABRIR_CORTINA","CERRAR_CORTINA","PAPER_W","DURACION_DE_SERVICIO"):
                    avg = round(base[c].fillna(0).mean(), 4)
                    row[c] = avg; gavg[c].append(avg)
                row["RECIBO_HRS"] = round(
                    row["ABRIR_CORTINA"]+row["CERRAR_CORTINA"]+row["PAPER_W"], 4)
                gavg["RECIBO_HRS"].append(row["RECIBO_HRS"])
                for c in ("LLEGADA_A_TRAFICO","SALIDA_DE_CD","LOS_HRS"):
                    vals = sb[c].dropna(); vals = vals[vals > 0]
                    avg  = round(vals.mean(), 4) if len(vals) else None
                    row[c] = avg
                    if avg is not None: gavg[c].append(avg)
        else:
            for pc in PUNT_COLS: row[pc] = ""

        if not se.empty:
            se_b = se[se["RECIBO_HRS"].notna() & (se["RECIBO_HRS"] > 0)]
            row["# CITAS_ED"] = len(se)
            row["RECIBO_ED_AVG"] = round(se_b["RECIBO_HRS"].mean(), 4) if not se_b.empty else None
            if row["RECIBO_ED_AVG"]: g_ed_hrs.append(row["RECIBO_ED_AVG"])
        else:
            row["# CITAS_ED"] = ""; row["RECIBO_ED_AVG"] = ""

        rows.append(row)

    grand = {
        "FECHA":"TOTAL GENERAL","# CITA":len(df_b),
        "VENDOR":"","CEDIS":"","CEDIS_NOMBRE":"",
        "TIPO_CITA":"","CITAS_CORRECTAS":"","SW":"",
        "PUNTUALIDAD":"","DIFERENCIA_MIN":"","_TIPO":"TOT","RECIBO_ED":None,
    }
    for c in NUM_BKHL:
        avgs = gavg[c]
        grand[c] = round(sum(avgs)/len(avgs), 4) if avgs else None
    if grand.get("ABRIR_CORTINA") is not None:
        grand["RECIBO_HRS"] = round(
            grand["ABRIR_CORTINA"]+grand["CERRAR_CORTINA"]+grand["PAPER_W"], 4)
    grand.update(_punt_counts(df_b, len(df_b)))
    grand["# CITAS_ED"]    = len(df_e)
    grand["RECIBO_ED_AVG"] = round(sum(g_ed_hrs)/len(g_ed_hrs),4) if g_ed_hrs else None
    rows.append(grand)

    return pd.DataFrame(rows, columns=COLS_DET)


# ─── Formato ──────────────────────────────────────────────────────────────────
def fmt_hrs(h) -> str:
    if h is None or h == "":
        return ""
    try:
        h = float(h)
    except (ValueError, TypeError):
        return ""
    if np.isnan(h) or h <= 0:
        return ""
    hh = int(h); mm = round((h - hh) * 60)
    if mm == 60: hh += 1; mm = 0
    return f"{hh}:{mm:02d}"


def _styles(hdr_hex: str, alt_hex: str, tot_hex: str) -> dict:
    thin = Side(style="thin", color="CCCCCC")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    return {
        "hdr": PatternFill("solid", fgColor=hdr_hex),
        "alt": PatternFill("solid", fgColor=alt_hex),
        "tot": PatternFill("solid", fgColor=tot_hex),
        "ed":  PatternFill("solid", fgColor="DAEEF3"),
        "emp": PatternFill("solid", fgColor="F5F5F5"),
        "grd": PatternFill("solid", fgColor="2E4057"),
        "wf":  Font(bold=True, color="FFFFFF", size=10),
        "hf":  Font(bold=True, color="FFFFFF", size=9),
        "bf":  Font(bold=True, size=9),
        "nf":  Font(size=9),
        "gf":  Font(bold=True, color="FFFFFF", size=9),
        "ef":  Font(italic=True, color="AAAAAA", size=9),
        "edf": Font(size=9, color="17375E"),
        "ctr": Alignment(horizontal="center", vertical="center"),
        "lft": Alignment(horizontal="left",   vertical="center"),
        "brd": brd,
    }

_P_GREEN  = PatternFill("solid", fgColor="E2EFDA")
_P_YELLOW = PatternFill("solid", fgColor="FFEB9C")
_P_RED    = PatternFill("solid", fgColor="FFC7CE")
_F_GREEN  = Font(color="375623", size=9)
_F_YELLOW = Font(color="9C5700", size=9)
_F_RED    = Font(color="9C0006", size=9)

HDR_LABELS = {c: c for c in COLS_DET}
HDR_LABELS["RECIBO_HRS"]    = "RECIBO\n(BKHL)"
HDR_LABELS["RECIBO_ED"]     = "RECIBO\n(ENT. DIRECTA)"
HDR_LABELS["RECIBO_ED_AVG"] = "RECIBO_ED\nPROM"


def write_sheet(ws, df_b: pd.DataFrame, df_e: pd.DataFrame,
                df_tot: pd.DataFrame, titulo: str,
                fecha_ini: str, fecha_fin: str,
                hdr_hex: str, alt_hex: str, tot_hex: str):
    ws.title = titulo[:31]
    s     = _styles(hdr_hex, alt_hex, tot_hex)
    ncols = len(COLS_DET)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(1, 1,
        f"BKHL + Ent. Directas - {titulo} | {fecha_ini} al {fecha_fin} | "
        f"{len(df_b):,} BKHL + {len(df_e):,} ED | tiempos hh:mm")
    t.fill = s["hdr"]; t.font = s["wf"]
    t.alignment = Alignment(horizontal="center", vertical="center")

    for j, col in enumerate(COLS_DET, 1):
        c = ws.cell(2, j, HDR_LABELS.get(col, col))
        c.fill = s["hdr"]; c.font = s["hf"]
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = s["brd"]

    df_b2 = df_b.copy(); df_b2["RECIBO_ED"] = np.nan
    df_e2 = df_e.copy()
    df_e2["RECIBO_ED"]  = df_e2["RECIBO_HRS"]
    df_e2["RECIBO_HRS"] = np.nan

    combined = pd.concat([df_b2, df_e2], ignore_index=True)
    combined["VENDOR"] = pd.Categorical(combined["VENDOR"],
                                        categories=VENDOR_NAMES, ordered=True)
    combined = combined.sort_values(["VENDOR","FECHA","CEDIS"]).reset_index(drop=True)

    row_i = 3; prev_vendor = None

    for _, row in combined.iterrows():
        vendor = row["VENDOR"]
        is_ed  = row["_TIPO"] == "ED"
        fill   = s["ed"] if is_ed else (s["alt"] if vendor != prev_vendor else None)
        prev_vendor = vendor

        for j, col in enumerate(COLS_DET, 1):
            val = row.get(col)
            if col in TOT_ONLY:
                cell_val = None
            elif col in NUM_BKHL or col == "RECIBO_ED":
                cell_val = fmt_hrs(val) or None
            else:
                cell_val = None if (
                    val is None or (isinstance(val, float) and np.isnan(val))
                ) else val

            dc = ws.cell(row_i, j, cell_val)
            if col == "PUNTUALIDAD" and cell_val:
                if cell_val == "ANTICIPADA":
                    dc.fill = _P_GREEN;  dc.font = _F_GREEN
                elif cell_val == "A TIEMPO":
                    dc.fill = _P_YELLOW; dc.font = _F_YELLOW
                elif cell_val == "DESPUES":
                    dc.fill = _P_RED;    dc.font = _F_RED
                else:
                    dc.font = s["nf"]
                    if fill: dc.fill = fill
            elif col == "RECIBO_ED" and cell_val:
                dc.fill = s["ed"]; dc.font = s["edf"]
            else:
                dc.font = s["bf"] if col in ("VENDOR","RECIBO_HRS","LOS_HRS") else s["nf"]
                if fill: dc.fill = fill
            dc.alignment = s["lft"] if col in LEFT_COLS else s["ctr"]
            dc.border = s["brd"]
        row_i += 1

    ws.row_dimensions[row_i].height = 5
    row_i += 1

    vendors_con_datos = set(combined["VENDOR"].dropna().unique())
    for _, row in df_tot.iterrows():
        is_grand = str(row.get("FECHA","")).startswith("TOTAL GENERAL")
        vendor   = str(row.get("VENDOR",""))
        is_empty = (not is_grand) and (vendor not in vendors_con_datos)

        fill = s["grd"] if is_grand else (s["emp"] if is_empty else s["tot"])
        font = s["gf"]  if is_grand else (s["ef"]  if is_empty else s["bf"])

        for j, col in enumerate(COLS_DET, 1):
            val = row.get(col)
            if col in NUM_BKHL:
                cell_val = fmt_hrs(val) or None
            elif col == "RECIBO_ED":
                # En filas de total: mostrar promedio ED en la misma columna
                cell_val = fmt_hrs(row.get("RECIBO_ED_AVG")) or None
            elif col == "RECIBO_ED_AVG":
                cell_val = None   # ya se muestra en RECIBO_ED
            else:
                cell_val = None if (
                    val is None or (isinstance(val, float) and np.isnan(val))
                ) else val
            dc = ws.cell(row_i, j, cell_val)
            dc.fill = fill; dc.font = font
            if col == "RECIBO_ED" and cell_val:
                dc.fill = s["ed"]; dc.font = s["edf"] if not is_grand else s["gf"]
            dc.alignment = s["lft"] if col in LEFT_COLS else s["ctr"]
            dc.border = s["brd"]
        row_i += 1

    for j, col in enumerate(COLS_DET, 1):
        ws.column_dimensions[get_column_letter(j)].width = COL_W.get(col, 12)
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}2"


# ─── Configuracion de hojas ───────────────────────────────────────────────────
SHEET_CFG = [
    ("CUAUTITLAN (N1+N2)", CEDIS_N1+CEDIS_N2, "833C00","FCE4D6","F4B183"),
    ("NAVE 1",             CEDIS_N1,           "7B3F00","FDE9D9","F4B183"),
    ("NAVE 2",             CEDIS_N2,           "4A1942","F5E6F5","D5A6E0"),
    ("NAVE 3 (DET 7492)",  CEDIS_N3,           "1F497D","DCE6F1","9DC3E6"),
]


def build_workbook(df_bkhl: pd.DataFrame, df_ed: pd.DataFrame,
                   out_file: Path, fecha_ini: str, fecha_fin: str):
    fi = pd.Timestamp(fecha_ini); ff = pd.Timestamp(fecha_fin)
    db = df_bkhl[(df_bkhl["FECHA"] >= fi) & (df_bkhl["FECHA"] <= ff)].copy()
    de = df_ed  [(df_ed  ["FECHA"] >= fi) & (df_ed  ["FECHA"] <= ff)].copy()

    wb = Workbook(); first = True
    for titulo, cedis, hdr, alt, tot in SHEET_CFG:
        sb     = filter_sheet(db, cedis)
        se     = filter_sheet(de, cedis)
        tot_df = make_totals(sb, se)
        ws     = wb.active if first else wb.create_sheet()
        first  = False
        write_sheet(ws, sb, se, tot_df, titulo, fecha_ini, fecha_fin, hdr, alt, tot)
        print(f"  {titulo}: {len(sb):,} BKHL | {len(se):,} ED")

    wb.save(out_file)
    print(f"  Guardado: {out_file.name}")


# ─── Main ──────────────────────────────────────────────────────────────────────
print(f"Consultando BKHL ({FECHA_INICIO_GLOBAL} al {FECHA_FIN_GLOBAL}) ...")
df_bkhl_all = query_bq(TIPOS_BKHL)
print(f"  {len(df_bkhl_all):,} registros BKHL raw")

print("Consultando Entregas Directas ...")
df_ed_all = query_bq(TIPOS_ED)
print(f"  {len(df_ed_all):,} registros ED raw")

df_bkhl_all = enrich(df_bkhl_all, "BKHL")
df_ed_all   = enrich(df_ed_all,   "ED")
print(f"  BKHL final: {len(df_bkhl_all):,} | ED final: {len(df_ed_all):,}")

print("\n[1/2] Generando JUNIO ...")
build_workbook(df_bkhl_all, df_ed_all,
               BASE / "top75_cuautitlan_junio_con_ed.xlsx",
               "2026-06-01", "2026-06-30")

print("\n[2/2] Generando JULIO 1-10 ...")
build_workbook(df_bkhl_all, df_ed_all,
               BASE / "top75_cuautitlan_jul1_10_con_ed.xlsx",
               "2026-07-01", "2026-07-10")

print("\nListo!")
print("  top75_cuautitlan_junio_con_ed.xlsx")
print("  top75_cuautitlan_jul1_10_con_ed.xlsx")
