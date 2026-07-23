"""
export_recibo_v2.py
Tiempo de recibo = AVG((ABRIR_CORTINA + CERRAR_CORTINA + PAPER_W) / 60)
- DIRECTO : TIPO_CITA Proveedor/Cita Nueva, CITAS_CORRECTAS = 1
- BKHL    : TIPO_CITA Backhaul/CNV Backhaul/Repro BKH, sin filtro CITAS_CORRECTAS
Estructura: Cuautitlan (N1+N2) | N1 | N2 | resto de CDs | Total
Fila extra: TOTAL por CD al final.
"""
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.cloud import bigquery

BASE      = Path(__file__).parent
BQ_TABLE  = "wmt-intl-supplychain-gcp-prod.MR101_WM_AD_HOC.SCH_YMS_SEMANAL"
FECHA_INI = "2026-07-01"
FECHA_FIN = "2026-07-10"

# ─── Vendors ─────────────────────────────────────────────────────────────────
VENDORS_KW = [
    "BONAFONT", "PEPSICO", "NIAGARA", "SUPREMA", "FRABEL",
    "HERDEZ", "JUGOS DEL VALLE", "KIMBERLY", "NESTLE",
    "MONDELEZ", "PROCTER", "SANTA CLARA", "UNILEVER",
]
VENDOR_DISPLAY = {
    "BONAFONT":        "BONAFONT SA CV",
    "PEPSICO":         "COMERC PEPSICO MEXICO S RL CV",
    "NIAGARA":         "EMBOTELLAD NIAGARA D MX S RLCV",
    "SUPREMA":         "ENVASADORA LA SUPREMA SA DE CV",
    "FRABEL":          "FRABEL SA DE CV",
    "HERDEZ":          "HERDEZ SA DE CV",
    "JUGOS DEL VALLE": "JUGOS DEL VALLE SAPI DE CV",
    "KIMBERLY":        "KIMBERLY CLARK MEXICO SA B CV",
    "NESTLE":          "MARCAS NESTLE SA CV",
    "MONDELEZ":        "MONDELEZ MEXICO S DE RL DE CV",
    "PROCTER":         "PROCTER AND GAMBLE MEXICO INC",
    "SANTA CLARA":     "SANTA CLARA MERC PACHU S RL CV",
    "UNILEVER":        "UNILEVER DE MEXICO S RL CV",
}
VENDOR_LABELS = list(VENDOR_DISPLAY.values())

# ─── CEDIS agrupados ─────────────────────────────────────────────────────────
# Cuautitlan general = N1 + N2 juntos
CUAU_ALL = [7494, 7464]
CUAU_N1  = [7494]
CUAU_N2  = [7464]

CD_SIMPLE = {
    "Megapark SMO":  [6388],
    "Santa Barbara": [7457, 7482],
    "Chihuahua":     [4640, 5780],
    "Culiacan":      [4971, 7455, 7487],
    "Mexicali":      [4924, 6140],
    "Monterrey":     [4995, 7461, 7490, 8806],
    "Chalco":        [7459, 7471, 7505],
    "Guadalajara":   [5907, 6238, 7460, 7493],
    "Merida":        [4188, 7103, 7506],
    "Villahermosa":  [6550, 7453, 7468],
}

# Orden final de columnas (sin Total, se agrega despues)
CD_COLS = ["Cuautitlan", "Cuautitlan N1", "Cuautitlan N2"] + list(CD_SIMPLE.keys())

# Mapa CEDIS -> lista de etiquetas de columna a la que pertenece
def build_cedis_map():
    m = {}
    for c in CUAU_ALL:
        m.setdefault(c, []).append("Cuautitlan")
    for c in CUAU_N1:
        m.setdefault(c, []).append("Cuautitlan N1")
    for c in CUAU_N2:
        m.setdefault(c, []).append("Cuautitlan N2")
    for cd_name, ids in CD_SIMPLE.items():
        for c in ids:
            m.setdefault(c, []).append(cd_name)
    return m

CEDIS_MAP = build_cedis_map()


def label_vendor(v: str):
    v = str(v).upper()
    for kw in VENDORS_KW:
        if kw in v:
            return VENDOR_DISPLAY[kw]
    return None


def query_raw(tipo_filter: str, correctas_filter: str) -> pd.DataFrame:
    client = bigquery.Client()
    q = f"""
    SELECT
        CEDIS,
        VENDOR,
        SAFE_CAST(ABRIR_CORTINA  AS FLOAT64) AS ABRIR,
        SAFE_CAST(CERRAR_CORTINA AS FLOAT64) AS CERRAR,
        SAFE_CAST(PAPER_W        AS FLOAT64) AS PAPER
    FROM `{BQ_TABLE}`
    WHERE ARRIVAL_DATE BETWEEN '{FECHA_INI}' AND '{FECHA_FIN}'
      AND {tipo_filter}
      {correctas_filter}
    """
    df = client.query(q).to_dataframe()
    df["CEDIS"] = pd.to_numeric(df["CEDIS"], errors="coerce")
    return df


def build_matrix(df: pd.DataFrame) -> pd.DataFrame:
    """
    Calcula promedio simple de (ABRIR + CERRAR + PAPER) / 60
    por VENDOR_LABEL x columna CD.
    """
    df = df.copy()
    df["VENDOR_LABEL"] = df["VENDOR"].apply(label_vendor)
    df = df[df["VENDOR_LABEL"].notna()].copy()

    # Metrica por fila
    df["ABRIR"]  = df["ABRIR"].fillna(0)
    df["CERRAR"] = df["CERRAR"].fillna(0)
    df["PAPER"]  = df["PAPER"].fillna(0)
    df["T_HRS"]  = (df["ABRIR"] + df["CERRAR"] + df["PAPER"]) / 60.0
    # Excluir filas con T_HRS = 0 (sin datos)
    df = df[df["T_HRS"] > 0].copy()

    # Expandir cada fila a sus columnas CD correspondientes
    rows = []
    for _, row in df.iterrows():
        cedis = int(row["CEDIS"]) if not pd.isna(row["CEDIS"]) else -1
        cols_dest = CEDIS_MAP.get(cedis, [])
        for cd_col in cols_dest:
            rows.append({"VENDOR_LABEL": row["VENDOR_LABEL"],
                         "CD_COL": cd_col,
                         "T_HRS": row["T_HRS"]})

    if not rows:
        return pd.DataFrame(index=VENDOR_LABELS, columns=CD_COLS + ["Total"], dtype=float)

    exp = pd.DataFrame(rows)
    # Promedio simple
    pivot = exp.groupby(["VENDOR_LABEL", "CD_COL"])["T_HRS"].mean().unstack("CD_COL")

    # Asegurar todas las columnas y filas
    mat = pd.DataFrame(index=VENDOR_LABELS, columns=CD_COLS, dtype=float)
    for v in VENDOR_LABELS:
        if v in pivot.index:
            for c in CD_COLS:
                if c in pivot.columns:
                    mat.loc[v, c] = pivot.loc[v, c]

    # Columna Total = promedio simple de todas las columnas con dato por vendor
    mat["Total"] = mat[CD_COLS].mean(axis=1, skipna=True)

    # Fila TOTAL = promedio simple de todos los vendors con dato por CD
    total_row = pd.Series(dtype=float)
    for c in CD_COLS + ["Total"]:
        col_data = mat[c].dropna()
        total_row[c] = col_data.mean() if len(col_data) else np.nan
    mat.loc["TOTAL GENERAL", :] = total_row

    return mat


def fmt_hrs(h):
    """Convierte horas decimales a hh:mm."""
    if h is None or (isinstance(h, float) and (np.isnan(h) or h <= 0)):
        return ""
    hh = int(h)
    mm = round((h - hh) * 60)
    if mm == 60:
        hh += 1; mm = 0
    return f"{hh}:{mm:02d}"


# ─── Excel writer ─────────────────────────────────────────────────────────────
ALL_COLS = CD_COLS + ["Total"]

def write_sheet(ws, mat: pd.DataFrame, titulo: str,
                hdr_hex: str, alt_hex: str, total_hex: str, cuau_hex: str):
    ws.title = titulo[:31]

    hdr_fill   = PatternFill("solid", fgColor=hdr_hex)
    alt_fill   = PatternFill("solid", fgColor=alt_hex)
    total_fill = PatternFill("solid", fgColor=total_hex)
    cuau_fill  = PatternFill("solid", fgColor=cuau_hex)
    wht_font   = Font(bold=True, color="FFFFFF", size=10)
    hdr_font   = Font(bold=True, color="FFFFFF", size=9)
    bold_f     = Font(bold=True, size=9)
    norm_f     = Font(size=9)
    center     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_a     = Alignment(horizontal="left",   vertical="center")
    thin       = Side(style="thin", color="CCCCCC")
    brd        = Border(left=thin, right=thin, top=thin, bottom=thin)

    ncols = len(ALL_COLS) + 1  # +1 para columna Proveedor

    # Fila 1: Titulo principal
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(1, 1, f"TIEMPO DE RECIBO  {titulo}  {FECHA_INI} al {FECHA_FIN}  (hh:mm, promedio simple)")
    t.fill = hdr_fill; t.font = wht_font; t.alignment = center

    # Fila 2: sub-cabecera Cuautitlan agrupada
    # Merge columnas Cuautitlan (col 2), N1 (col 3), N2 (col 4)
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=4)
    cc = ws.cell(2, 2, "CUAUTITLAN")
    cc.fill = cuau_fill; cc.font = hdr_font; cc.alignment = center; cc.border = brd
    ws.cell(2, 1, "").fill = hdr_fill  # esquina

    # Resto de columnas en fila 2 (vacias para las que no son Cuautitlan)
    for j in range(5, ncols + 1):
        ws.cell(2, j, "").fill = hdr_fill

    # Fila 3: Proveedor + nombres de columnas
    ws.cell(3, 1, "Proveedor").fill = hdr_fill
    ws.cell(3, 1).font = hdr_font; ws.cell(3, 1).alignment = center
    ws.cell(3, 1).border = brd

    COL_NAMES = ["Cuautitlan", "N1", "N2",
                 "Megapark SMO", "Santa Barbara", "Chihuahua",
                 "Culiacan", "Mexicali", "Monterrey", "Chalco",
                 "Guadalajara", "Merida", "Villahermosa", "Total"]
    for j, name in enumerate(COL_NAMES, 2):
        is_cuau = j in (2, 3, 4)
        fill = cuau_fill if is_cuau else (hdr_fill if name != "Total" else
               PatternFill("solid", fgColor="2E4057"))
        c = ws.cell(3, j, name)
        c.fill = fill; c.font = hdr_font; c.alignment = center; c.border = brd

    # Datos
    all_row_labels = list(mat.index)
    for i, vendor in enumerate(all_row_labels, 4):
        is_total_row = vendor == "TOTAL GENERAL"
        rf = total_fill if is_total_row else (alt_fill if (i % 2 == 0) else None)

        vc = ws.cell(i, 1, vendor)
        vc.font = bold_f; vc.alignment = left_a; vc.border = brd
        if rf: vc.fill = rf

        for j, col in enumerate(ALL_COLS, 2):
            val = mat.loc[vendor, col]
            txt = fmt_hrs(val) if not (isinstance(val, float) and np.isnan(val)) else ""
            is_cuau_col = j in (2, 3, 4)
            is_total_col = col == "Total"
            dc = ws.cell(i, j, txt if txt else None)
            dc.font = bold_f if (is_total_row or is_total_col) else norm_f
            dc.alignment = center; dc.border = brd
            if is_total_row:
                dc.fill = total_fill
            elif is_cuau_col and rf:
                dc.fill = rf
            elif is_cuau_col:
                dc.fill = PatternFill("solid", fgColor="EBF3FB")
            elif rf:
                dc.fill = rf

    # Anchos
    ws.column_dimensions["A"].width = 36
    col_w = [14, 11, 11, 13, 13, 11, 11, 10, 11, 10, 13, 10, 13, 11]
    for j, w in enumerate(col_w, 2):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 32
    ws.freeze_panes = "B4"


# ─── Main ─────────────────────────────────────────────────────────────────────
print(f"BigQuery: {FECHA_INI} al {FECHA_FIN}")

print("  Directo (Proveedor/Cita Nueva, CITAS_CORRECTAS=1)...")
df_dir = query_raw(
    "UPPER(TRIM(TIPO_CITA)) IN ('PROVEEDOR', 'CITA NUEVA')",
    "AND CITAS_CORRECTAS = 1"
)
mat_dir = build_matrix(df_dir)
print(f"    {len(df_dir):,} filas raw | {mat_dir.iloc[:-1].notna().sum().sum()} celdas con dato")

print("  BKHL (Backhaul, CORRECTO e INCORRECTO)...")
df_bkhl = query_raw(
    "UPPER(TRIM(TIPO_CITA)) IN ('BACKHAUL', 'CNV BACKHAUL', 'REPRO BKH')",
    ""  # sin filtro CITAS_CORRECTAS
)
mat_bkhl = build_matrix(df_bkhl)
print(f"    {len(df_bkhl):,} filas raw | {mat_bkhl.iloc[:-1].notna().sum().sum()} celdas con dato")

out = BASE / "tiempos_recibo_v2.xlsx"
print(f"\nEscribiendo {out}...")
wb = Workbook()
write_sheet(
    wb.active, mat_dir, "DIRECTO",
    hdr_hex="1F4E79", alt_hex="D9E1F2",
    total_hex="BDD7EE", cuau_hex="2E75B6"
)
write_sheet(
    wb.create_sheet(), mat_bkhl, "BKHL",
    hdr_hex="833C00", alt_hex="FCE4D6",
    total_hex="F4B183", cuau_hex="C55A11"
)
wb.save(out)
print(f"Listo: {out}")
