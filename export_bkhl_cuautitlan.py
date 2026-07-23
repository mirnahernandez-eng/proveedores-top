"""
export_bkhl_cuautitlan.py
Registros BKHL (cita por cita) desde LLEGADA hasta SALIDA
para Cuautitlan (N1 + N2 totales) — 2026 completo.
Hojas: CUAUTITLAN (N1+N2), N1, N2
Totales por vendor al final de cada hoja.
"""
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.cloud import bigquery

BASE     = Path(__file__).parent
BQ_TABLE = "wmt-intl-supplychain-gcp-prod.MR101_WM_AD_HOC.SCH_YMS_SEMANAL"
FECHA_INI = "2026-06-01"
FECHA_FIN = "2026-06-30"

CUAU_N1 = [7494]
CUAU_N2 = [7464]
CUAU_ALL = CUAU_N1 + CUAU_N2

VENDORS_KW = [
    "BONAFONT", "PEPSICO", "NIAGARA", "SUPREMA", "FRABEL",
    "HERDEZ", "JUGOS DEL VALLE", "KIMBERLY", "NESTLE",
    "MONDELEZ", "PROCTER", "SANTA CLARA", "UNILEVER",
]
VENDOR_DISPLAY = {
    "BONAFONT":        "BONAFONT SA CV",
    "PEPSICO":         "COMERC PEPSICO MEXICO S RL CV",
    "NIAGARA":         "EMBOTELLAD NIAGARA D MX S RLCV",
    "SUPREMA":         "ENVASADORA LA SUPREMA SA DE CV",
    "FRABEL":          "FRABEL SA DE CV",
    "HERDEZ":          "HERDEZ SA DE CV",
    "JUGOS DEL VALLE": "JUGOS DEL VALLE SAPI DE CV",
    "KIMBERLY":        "KIMBERLY CLARK MEXICO SA B CV",
    "NESTLE":          "MARCAS NESTLE SA CV",
    "MONDELEZ":        "MONDELEZ MEXICO S DE RL DE CV",
    "PROCTER":         "PROCTER AND GAMBLE MEXICO INC",
    "SANTA CLARA":     "SANTA CLARA MERC PACHU S RL CV",
    "UNILEVER":        "UNILEVER DE MEXICO S RL CV",
}
VENDOR_ORDER = list(VENDOR_DISPLAY.values())

CEDIS_LABEL = {7494: "Cuautitlan N1", 7464: "Cuautitlan N2"}

# Columnas numericas para totales
NUM_COLS = ["LLEGADA_A_TRAFICO", "ABRIR_CORTINA", "CERRAR_CORTINA",
            "PAPER_W", "SALIDA_DE_CD", "DURACION_DE_SERVICIO",
            "RECIBO_HRS", "LOS_HRS"]

COLS_DISPLAY = [
    "FECHA", "# CITA", "VENDOR", "CEDIS", "CEDIS_NOMBRE",
    "TIPO_CITA", "CITAS_CORRECTAS", "SW",
    "LLEGADA_A_TRAFICO", "ABRIR_CORTINA", "CERRAR_CORTINA",
    "PAPER_W", "SALIDA_DE_CD", "DURACION_DE_SERVICIO",
    "RECIBO_HRS", "LOS_HRS",
]

COL_WIDTHS = {
    "FECHA": 13, "# CITA": 16, "VENDOR": 34, "CEDIS": 8,
    "CEDIS_NOMBRE": 18, "TIPO_CITA": 16, "CITAS_CORRECTAS": 10, "SW": 6,
    "LLEGADA_A_TRAFICO": 14, "ABRIR_CORTINA": 13, "CERRAR_CORTINA": 14,
    "PAPER_W": 10, "SALIDA_DE_CD": 13, "DURACION_DE_SERVICIO": 18,
    "RECIBO_HRS": 12, "LOS_HRS": 12,
}


def label_vendor(v: str):
    v = str(v).upper()
    for kw in VENDORS_KW:
        if kw in v:
            return VENDOR_DISPLAY[kw]
    return None


def query_bkhl() -> pd.DataFrame:
    client = bigquery.Client()
    cedis_str = ",".join(str(c) for c in CUAU_ALL)
    q = f"""
    SELECT
        ARRIVAL_DATE                                   AS FECHA,
        APPOINTMENT_NBR                                AS `# CITA`,
        VENDOR,
        SAFE_CAST(CEDIS AS INT64)                      AS CEDIS,
        NOMBRE_CEDIS                                   AS CEDIS_NOMBRE,
        TIPO_CITA,
        CITAS_CORRECTAS,
        SW,
        ROUND(SAFE_CAST(LLEGADA_A_TRAFICO    AS FLOAT64) / 60.0, 4) AS LLEGADA_A_TRAFICO,
        ROUND(SAFE_CAST(ABRIR_CORTINA        AS FLOAT64) / 60.0, 4) AS ABRIR_CORTINA,
        ROUND(SAFE_CAST(CERRAR_CORTINA       AS FLOAT64) / 60.0, 4) AS CERRAR_CORTINA,
        ROUND(SAFE_CAST(PAPER_W              AS FLOAT64) / 60.0, 4) AS PAPER_W,
        ROUND(SAFE_CAST(SALIDA_DE_CD         AS FLOAT64) / 60.0, 4) AS SALIDA_DE_CD,
        ROUND(SAFE_CAST(DURACION_DE_SERVICIO AS FLOAT64) / 60.0, 4) AS DURACION_DE_SERVICIO,
        ROUND(
            (COALESCE(SAFE_CAST(ABRIR_CORTINA  AS FLOAT64), 0) +
             COALESCE(SAFE_CAST(CERRAR_CORTINA AS FLOAT64), 0) +
             COALESCE(SAFE_CAST(PAPER_W        AS FLOAT64), 0)) / 60.0
        , 4) AS RECIBO_HRS,
        ROUND((
            COALESCE(SAFE_CAST(LLEGADA_A_TRAFICO    AS FLOAT64), 0) +
            COALESCE(SAFE_CAST(DURACION_DE_SERVICIO AS FLOAT64), 0) +
            COALESCE(SAFE_CAST(SALIDA_DE_CD         AS FLOAT64), 0)
        ) / 60.0, 4) AS LOS_HRS
    FROM `{BQ_TABLE}`
    WHERE ARRIVAL_DATE BETWEEN '{FECHA_INI}' AND '{FECHA_FIN}'
      AND UPPER(TRIM(TIPO_CITA)) IN ('BACKHAUL', 'CNV BACKHAUL', 'REPRO BKH')
      AND SAFE_CAST(CEDIS AS INT64) IN ({cedis_str})
      AND (
            COALESCE(SAFE_CAST(ABRIR_CORTINA  AS FLOAT64), 0) +
            COALESCE(SAFE_CAST(CERRAR_CORTINA AS FLOAT64), 0) +
            COALESCE(SAFE_CAST(PAPER_W        AS FLOAT64), 0) +
            COALESCE(SAFE_CAST(LLEGADA_A_TRAFICO    AS FLOAT64), 0) +
            COALESCE(SAFE_CAST(DURACION_DE_SERVICIO AS FLOAT64), 0) +
            COALESCE(SAFE_CAST(SALIDA_DE_CD         AS FLOAT64), 0)
          ) > 0
    ORDER BY FECHA, CEDIS, VENDOR
    """
    return client.query(q).to_dataframe()


def enrich(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["VENDOR"] = df["VENDOR"].apply(
        lambda v: label_vendor(str(v)) if label_vendor(str(v)) else str(v)
    )
    # Solo vendors objetivo
    df = df[df["VENDOR"].isin(VENDOR_ORDER)].copy()
    df["VENDOR"] = pd.Categorical(df["VENDOR"], categories=VENDOR_ORDER, ordered=True)
    df = df.sort_values(["VENDOR", "FECHA", "CEDIS"]).reset_index(drop=True)
    return df


def make_totals(df: pd.DataFrame) -> pd.DataFrame:
    """Totales por vendor + TOTAL GENERAL como promedio de promedios.
    RECIBO_HRS se deriva de ABRIR+CERRAR+PAPER para garantizar que sumen igual.
    """
    rows = []
    vendor_avgs = {col: [] for col in NUM_COLS}

    for vendor in VENDOR_ORDER:
        sub = df[df["VENDOR"] == vendor]
        if sub.empty:
            continue

        # Base: registros con algun tiempo de andén > 0
        mask = sub["RECIBO_HRS"].notna() & (sub["RECIBO_HRS"] > 0)
        base = sub.loc[mask]
        if base.empty:
            continue

        row = {"FECHA": f"TOTAL — {vendor}", "# CITA": len(sub),
               "VENDOR": vendor, "CEDIS": "", "CEDIS_NOMBRE": "",
               "TIPO_CITA": "", "CITAS_CORRECTAS": "", "SW": ""}

        # Columnas de andén: fillna(0) sobre misma base
        for col in ("ABRIR_CORTINA", "CERRAR_CORTINA", "PAPER_W", "DURACION_DE_SERVICIO"):
            avg = round(base[col].fillna(0).mean(), 4)
            row[col] = avg
            vendor_avgs[col].append(avg)

        # RECIBO_HRS = suma de los tres componentes (garantiza consistencia)
        recibo = round(
            row["ABRIR_CORTINA"] + row["CERRAR_CORTINA"] + row["PAPER_W"], 4
        )
        row["RECIBO_HRS"] = recibo
        vendor_avgs["RECIBO_HRS"].append(recibo)

        # Resto de columnas de tiempo (LLEGADA, SALIDA, LOS)
        for col in ("LLEGADA_A_TRAFICO", "SALIDA_DE_CD", "LOS_HRS"):
            vals = sub[col].dropna()
            vals = vals[vals > 0]
            avg = round(vals.mean(), 4) if len(vals) else None
            row[col] = avg
            if avg is not None:
                vendor_avgs[col].append(avg)

        rows.append(row)

    # TOTAL GENERAL = promedio de promedios por vendor
    grand = {"FECHA": "TOTAL GENERAL", "# CITA": len(df),
             "VENDOR": "", "CEDIS": "", "CEDIS_NOMBRE": "",
             "TIPO_CITA": "", "CITAS_CORRECTAS": "", "SW": ""}
    for col in NUM_COLS:
        avgs = vendor_avgs[col]
        grand[col] = round(sum(avgs) / len(avgs), 4) if avgs else None
    # Garantizar que RECIBO_HRS grand = ABRIR_grand + CERRAR_grand + PAPER_grand
    if grand["ABRIR_CORTINA"] is not None:
        grand["RECIBO_HRS"] = round(
            grand["ABRIR_CORTINA"] + grand["CERRAR_CORTINA"] + grand["PAPER_W"], 4
        )
    rows.append(grand)

    return pd.DataFrame(rows, columns=COLS_DISPLAY)

    return pd.DataFrame(rows, columns=COLS_DISPLAY)


def fmt_hrs(h) -> str:
    """Convierte horas decimales a hh:mm. Devuelve '' si no hay dato."""
    if h is None or (isinstance(h, float) and (np.isnan(h) or h <= 0)):
        return ""
    hh = int(h)
    mm = round((h - hh) * 60)
    if mm == 60:
        hh += 1; mm = 0
    return f"{hh}:{mm:02d}"


# ─── Estilos ──────────────────────────────────────────────────────────────────
def _styles(hdr_hex: str, alt_hex: str, tot_hex: str):
    return {
        "hdr":  PatternFill("solid", fgColor=hdr_hex),
        "alt":  PatternFill("solid", fgColor=alt_hex),
        "tot":  PatternFill("solid", fgColor=tot_hex),
        "grand": PatternFill("solid", fgColor="2E4057"),
        "wht_f": Font(bold=True, color="FFFFFF", size=10),
        "hdr_f": Font(bold=True, color="FFFFFF", size=9),
        "bold_f": Font(bold=True, size=9),
        "norm_f": Font(size=9),
        "grand_f": Font(bold=True, color="FFFFFF", size=9),
        "center": Alignment(horizontal="center", vertical="center"),
        "left":   Alignment(horizontal="left",   vertical="center"),
        "brd": Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        ),
    }


def write_sheet(ws, df_data: pd.DataFrame, df_tot: pd.DataFrame,
                titulo: str, hdr_hex: str, alt_hex: str, tot_hex: str,
                fecha_ini: str, fecha_fin: str):
    ws.title = titulo[:31]
    s = _styles(hdr_hex, alt_hex, tot_hex)
    cols = COLS_DISPLAY
    ncols = len(cols)
    LEFT_COLS = {"FECHA", "VENDOR", "CEDIS_NOMBRE", "TIPO_CITA"}

    # ── Titulo ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(1, 1,
        f"BKHL — {titulo} | {fecha_ini} al {fecha_fin} | {len(df_data):,} registros | tiempos en HORAS")
    t.fill = s["hdr"]; t.font = s["wht_f"]; t.alignment = s["center"]

    # ── Cabecera ──
    for j, col in enumerate(cols, 1):
        c = ws.cell(2, j, col)
        c.fill = s["hdr"]; c.font = s["hdr_f"]
        c.alignment = s["center"]; c.border = s["brd"]

    # ── Datos ──
    row_i = 3
    prev_vendor = None
    for _, row in df_data.iterrows():
        vendor = row["VENDOR"]
        rf = s["alt"] if (vendor != prev_vendor) else None
        prev_vendor = vendor
        for j, col in enumerate(cols, 1):
            val = row.get(col)
            if col in NUM_COLS:
                cell_val = fmt_hrs(val) or None
            else:
                cell_val = None if (val is None or (isinstance(val, float) and np.isnan(val))) else val
            dc = ws.cell(row_i, j, cell_val)
            dc.font = s["bold_f"] if col in ("VENDOR", "RECIBO_HRS", "LOS_HRS") else s["norm_f"]
            dc.alignment = s["left"] if col in LEFT_COLS else s["center"]
            dc.border = s["brd"]
            if rf:
                dc.fill = rf
        row_i += 1

    # ── Totales ──
    # Linea separadora
    ws.row_dimensions[row_i].height = 6
    row_i += 1

    for _, row in df_tot.iterrows():
        is_grand = str(row.get("FECHA", "")).startswith("TOTAL GENERAL")
        fill = s["grand"] if is_grand else s["tot"]
        font_f = s["grand_f"] if is_grand else s["bold_f"]
        for j, col in enumerate(cols, 1):
            val = row.get(col)
            if col in NUM_COLS:
                cell_val = fmt_hrs(val) or None
            else:
                cell_val = None if (val is None or (isinstance(val, float) and np.isnan(val))) else val
            dc = ws.cell(row_i, j, cell_val)
            dc.fill = fill; dc.font = font_f
            dc.alignment = s["left"] if col in LEFT_COLS else s["center"]
            dc.border = s["brd"]
        row_i += 1

    # ── Anchos ──
    for j, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = COL_WIDTHS.get(col, 12)
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}2"


# ─── Main ─────────────────────────────────────────────────────────────────────
print(f"Consultando BigQuery BKHL Cuautitlan: {FECHA_INI} al {FECHA_FIN} ...")
df_raw = query_bkhl()
print(f"  {len(df_raw):,} registros raw")

df = enrich(df_raw)
print(f"  {len(df):,} registros con vendors objetivo")

df_n1 = df[df["CEDIS"] == 7494].copy()
df_n2 = df[df["CEDIS"] == 7464].copy()
print(f"  N1: {len(df_n1):,}  N2: {len(df_n2):,}")

tot_all = make_totals(df)
tot_n1  = make_totals(df_n1)
tot_n2  = make_totals(df_n2)

out = BASE / "bkhl_cuau_junio_v2.xlsx"
print(f"\nEscribiendo {out} ...")
wb = Workbook()

write_sheet(wb.active,         df,    tot_all, "CUAUTITLAN (N1+N2)",
            "833C00", "FCE4D6", "F4B183", FECHA_INI, FECHA_FIN)
write_sheet(wb.create_sheet(), df_n1, tot_n1,  "NAVE 1",
            "7B3F00", "FDE9D9", "F4B183", FECHA_INI, FECHA_FIN)
write_sheet(wb.create_sheet(), df_n2, tot_n2,  "NAVE 2",
            "4A1942", "F5E6F5", "D5A6E0", FECHA_INI, FECHA_FIN)

wb.save(out)
print(f"Listo: {out}")
print(f"  CUAUTITLAN total : {len(df):,} citas")
print(f"  NAVE 1           : {len(df_n1):,} citas")
print(f"  NAVE 2           : {len(df_n2):,} citas")
