"""
export_niagara_bae_ingreso_julio.py
NIAGARA - TODOS LOS BAE - Julio 2026
Ordenado por LLEGADA_A_TRAFICO desc (tiempos de ingreso mas altos primero)
"""
import warnings; warnings.filterwarnings("ignore")
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.cloud import bigquery

BASE     = Path(__file__).parent
BQ_TABLE = "wmt-intl-supplychain-gcp-prod.MR101_WM_AD_HOC.SCH_YMS_SEMANAL"
OUT_FILE = BASE / "niagara_bae_ingreso_alto_julio.xlsx"

q = f"""
SELECT
    ARRIVAL_DATE                                             AS FECHA,
    APPOINTMENT_NBR                                          AS CITA,
    VENDOR,
    SAFE_CAST(CEDIS AS INT64)                                AS CEDIS,
    NOMBRE_CEDIS,
    TIPO_CITA,
    CITAS_CORRECTAS,
    SW,
    COALESCE(CITA_VS_LLEGADA, 'SIN DATO')                   AS PUNTUALIDAD,
    ROUND(SAFE_CAST(DIFERENCIA       AS FLOAT64), 2)        AS DIFERENCIA_MIN,
    ROUND(SAFE_CAST(LLEGADA_A_TRAFICO    AS FLOAT64)/60, 4) AS LLEGADA_HRS,
    ROUND(SAFE_CAST(ABRIR_CORTINA        AS FLOAT64)/60, 4) AS ABRIR_CORTINA,
    ROUND(SAFE_CAST(CERRAR_CORTINA       AS FLOAT64)/60, 4) AS CERRAR_CORTINA,
    ROUND(SAFE_CAST(PAPER_W              AS FLOAT64)/60, 4) AS PAPER_W,
    ROUND(SAFE_CAST(SALIDA_DE_CD         AS FLOAT64)/60, 4) AS SALIDA_HRS,
    ROUND(SAFE_CAST(DURACION_DE_SERVICIO AS FLOAT64)/60, 4) AS SERVICIO_HRS,
    ROUND((
        COALESCE(SAFE_CAST(ABRIR_CORTINA  AS FLOAT64),0)+
        COALESCE(SAFE_CAST(CERRAR_CORTINA AS FLOAT64),0)+
        COALESCE(SAFE_CAST(PAPER_W        AS FLOAT64),0)
    )/60, 4) AS RECIBO_HRS,
    ROUND((
        COALESCE(SAFE_CAST(LLEGADA_A_TRAFICO    AS FLOAT64),0)+
        COALESCE(SAFE_CAST(DURACION_DE_SERVICIO AS FLOAT64),0)+
        COALESCE(SAFE_CAST(SALIDA_DE_CD         AS FLOAT64),0)
    )/60, 4) AS LOS_HRS
FROM `{BQ_TABLE}`
WHERE ARRIVAL_DATE BETWEEN '2026-07-01' AND '2026-07-31'
  AND UPPER(VENDOR) LIKE '%NIAGARA%'
  AND UPPER(NOMBRE_CEDIS) LIKE '%BAE%'
  AND SAFE_CAST(LLEGADA_A_TRAFICO AS FLOAT64) > 0
ORDER BY LLEGADA_HRS DESC
"""

print("Consultando Niagara BAE - julio 2026 ...")
df = bigquery.Client().query(q).to_dataframe()
print(f"  {len(df):,} citas con ingreso registrado")

# ─── Helpers ──────────────────────────────────────────────────────────────────
HRS_COLS  = ["LLEGADA_HRS","ABRIR_CORTINA","CERRAR_CORTINA",
             "PAPER_W","SALIDA_HRS","SERVICIO_HRS","RECIBO_HRS","LOS_HRS"]
COLS      = ["FECHA","CITA","VENDOR","CEDIS","NOMBRE_CEDIS","TIPO_CITA",
             "CITAS_CORRECTAS","SW","PUNTUALIDAD","DIFERENCIA_MIN",
             "LLEGADA_HRS","ABRIR_CORTINA","CERRAR_CORTINA","PAPER_W",
             "SALIDA_HRS","SERVICIO_HRS","RECIBO_HRS","LOS_HRS"]
COL_W     = {
    "FECHA":13,"CITA":16,"VENDOR":32,"CEDIS":8,"NOMBRE_CEDIS":20,
    "TIPO_CITA":16,"CITAS_CORRECTAS":10,"SW":6,
    "PUNTUALIDAD":18,"DIFERENCIA_MIN":14,
    "LLEGADA_HRS":14,"ABRIR_CORTINA":13,"CERRAR_CORTINA":14,
    "PAPER_W":10,"SALIDA_HRS":13,"SERVICIO_HRS":14,
    "RECIBO_HRS":13,"LOS_HRS":12,
}
LEFT_COLS = {"FECHA","VENDOR","NOMBRE_CEDIS","TIPO_CITA","PUNTUALIDAD"}
_ANTICIP  = {"1 DIA ANTES","12-24 HORAS ANTES","6-12 HORAS ANTES","1-6 HORAS ANTES"}
thin      = Side(style="thin", color="CCCCCC")
BRD       = Border(left=thin, right=thin, top=thin, bottom=thin)
HDR_HEX   = "7B3F00"   # cafe BAE


def fmt_hrs(h) -> str:
    if h is None or h == "": return ""
    try:   h = float(h)
    except: return ""
    if np.isnan(h) or h <= 0: return ""
    hh = int(h); mm = round((h - hh) * 60)
    if mm == 60: hh += 1; mm = 0
    return f"{hh}:{mm:02d}"


def punt_style(v):
    v = str(v).upper().strip() if v else ""
    if v in _ANTICIP:
        return PatternFill("solid", fgColor="E2EFDA"), Font(color="375623", size=9)
    if v == "A TIEMPO":
        return PatternFill("solid", fgColor="FFEB9C"), Font(color="9C5700", size=9)
    if "DESPU" in v or "TARDE" in v:
        return PatternFill("solid", fgColor="FFC7CE"), Font(color="9C0006", size=9)
    return None, Font(size=9)


# Umbrales de ingreso
sal_vals = df["LLEGADA_HRS"].dropna(); sal_vals = sal_vals[sal_vals > 0]
avg_lleg = sal_vals.mean() if len(sal_vals) else 0
p75_lleg = sal_vals.quantile(0.75) if len(sal_vals) else 0
print(f"  Promedio LLEGADA: {fmt_hrs(avg_lleg)} | P75: {fmt_hrs(p75_lleg)}")

RED_FILL = PatternFill("solid", fgColor="FFE0E0")
ORG_FILL = PatternFill("solid", fgColor="FFF2CC")
ALT_FILL = PatternFill("solid", fgColor="FDE9D9")

# ─── Excel ────────────────────────────────────────────────────────────────────
wb = Workbook(); ws = wb.active
ws.title = "NIAGARA BAE JULIO"
ncols    = len(COLS)
hdr_fill = PatternFill("solid", fgColor=HDR_HEX)
tot_fill = PatternFill("solid", fgColor="2E4057")

# Titulo
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
t = ws.cell(1, 1,
    f"NIAGARA - TODOS LOS BAE | Julio 2026 | {len(df):,} citas | "
    f"Prom ingreso: {fmt_hrs(avg_lleg)} | P75: {fmt_hrs(p75_lleg)} | ordenado por LLEGADA desc")
t.fill = hdr_fill; t.font = Font(bold=True, color="FFFFFF", size=10)
t.alignment = Alignment(horizontal="center", vertical="center")

# Cabecera
HDR_LABELS = {c: c for c in COLS}
HDR_LABELS["LLEGADA_HRS"] = "LLEGADA_HRS\n(ingreso desc)"
for j, col in enumerate(COLS, 1):
    c = ws.cell(2, j, HDR_LABELS.get(col, col))
    c.fill = hdr_fill; c.font = Font(bold=True, color="FFFFFF", size=9)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BRD

# Datos
for i, (_, row) in enumerate(df.iterrows()):
    ri     = i + 3
    lleg_v = float(row.get("LLEGADA_HRS") or 0)
    row_fill = RED_FILL if lleg_v >= p75_lleg and lleg_v > 0 else (
               ORG_FILL if lleg_v >= avg_lleg and lleg_v > 0 else (
               ALT_FILL if i % 2 == 0 else None))

    for j, col in enumerate(COLS, 1):
        val      = row.get(col)
        cell_val = (fmt_hrs(val) or None) if col in HRS_COLS else (
            None if (val is None or (isinstance(val, float) and np.isnan(val))) else val)

        dc = ws.cell(ri, j, cell_val)

        if col == "PUNTUALIDAD":
            pf, pfont = punt_style(val)
            if pf: dc.fill = pf
            dc.font = pfont
        elif col == "LLEGADA_HRS":
            dc.fill = RED_FILL if lleg_v >= p75_lleg and lleg_v > 0 else (ORG_FILL if lleg_v > 0 else (row_fill or PatternFill()))
            dc.font = Font(bold=True, size=9,
                           color="C00000" if lleg_v >= p75_lleg and lleg_v > 0 else ("7F6000" if lleg_v > 0 else "000000"))
        else:
            dc.font = Font(bold=(col in ("CITA","CEDIS","RECIBO_HRS","LOS_HRS")), size=9)
            if row_fill: dc.fill = row_fill

        dc.alignment = Alignment(
            horizontal="left" if col in LEFT_COLS else "center",
            vertical="center")
        dc.border = BRD

# Separador + promedio
ws.row_dimensions[len(df) + 3].height = 5
ri_tot = len(df) + 4
ws.cell(ri_tot, 1, f"PROMEDIO ({len(df)} citas)").fill = tot_fill
ws.cell(ri_tot, 1).font = Font(bold=True, color="FFFFFF", size=9)
ws.cell(ri_tot, 1).alignment = Alignment(horizontal="center", vertical="center")
ws.cell(ri_tot, 1).border = BRD

for j, col in enumerate(COLS, 1):
    if j == 1: continue
    if col in HRS_COLS:
        vals = df[col].dropna(); vals = vals[vals > 0]
        cell_val = fmt_hrs(vals.mean()) if len(vals) else None
    else:
        cell_val = None
    dc = ws.cell(ri_tot, j, cell_val)
    dc.fill = tot_fill; dc.font = Font(bold=True, color="FFFFFF", size=9)
    dc.alignment = Alignment(horizontal="center", vertical="center")
    dc.border = BRD

# Leyenda
lr = ri_tot + 2
for txt, fill, desc in [
    ("ROJO:",     RED_FILL, f"Ingreso >= P75 ({fmt_hrs(p75_lleg)})"),
    ("AMARILLO:", ORG_FILL, f"Ingreso >= Promedio ({fmt_hrs(avg_lleg)})"),
]:
    ws.cell(lr, 1, txt).font  = Font(bold=True, size=9)
    ws.cell(lr, 1).fill       = fill
    ws.cell(lr, 2, desc).font = Font(size=9)
    lr += 1

# Anchos y freeze
for j, col in enumerate(COLS, 1):
    ws.column_dimensions[get_column_letter(j)].width = COL_W.get(col, 12)
ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 30
ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}2"

wb.save(OUT_FILE)
print(f"\nListo -> {OUT_FILE.name}")
print(f"  Citas con ingreso >= P75 ({fmt_hrs(p75_lleg)}): {int((df['LLEGADA_HRS'] >= p75_lleg).sum())}")
print(f"  Citas con ingreso >= prom ({fmt_hrs(avg_lleg)}): {int((df['LLEGADA_HRS'] >= avg_lleg).sum())}")
