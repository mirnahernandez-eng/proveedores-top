"""
export_bonafont_chalco_julio.py
Bonafont - CHALCO (CEDIS 7471) - Julio 2026
Ordenado por SALIDA_DE_CD desc (tiempos mas altos primero)
"""
import warnings; warnings.filterwarnings("ignore")
import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.cloud import bigquery

BASE     = Path(__file__).parent
BQ_TABLE = "wmt-intl-supplychain-gcp-prod.MR101_WM_AD_HOC.SCH_YMS_SEMANAL"
CEDIS    = 7471
OUT_FILE = BASE / "bonafont_chalco_julio_salida_alta.xlsx"

# ─── Query ────────────────────────────────────────────────────────────────────
q = f"""
SELECT
    ARRIVAL_DATE                                          AS FECHA,
    APPOINTMENT_NBR                                       AS CITA,
    VENDOR,
    SAFE_CAST(CEDIS AS INT64)                             AS CEDIS,
    NOMBRE_CEDIS,
    TIPO_CITA,
    CITAS_CORRECTAS,
    SW,
    COALESCE(CITA_VS_LLEGADA, 'SIN DATO')                AS PUNTUALIDAD,
    ROUND(SAFE_CAST(DIFERENCIA AS FLOAT64), 2)           AS DIFERENCIA_MIN,
    ROUND(SAFE_CAST(LLEGADA_A_TRAFICO    AS FLOAT64)/60, 4) AS LLEGADA_HRS,
    ROUND(SAFE_CAST(ABRIR_CORTINA        AS FLOAT64)/60, 4) AS ABRIR_CORTINA,
    ROUND(SAFE_CAST(CERRAR_CORTINA       AS FLOAT64)/60, 4) AS CERRAR_CORTINA,
    ROUND(SAFE_CAST(PAPER_W              AS FLOAT64)/60, 4) AS PAPER_W,
    ROUND(SAFE_CAST(SALIDA_DE_CD         AS FLOAT64)/60, 4) AS SALIDA_HRS,
    ROUND(SAFE_CAST(DURACION_DE_SERVICIO AS FLOAT64)/60, 4) AS SERVICIO_HRS,
    ROUND((
        COALESCE(SAFE_CAST(ABRIR_CORTINA  AS FLOAT64),0)+
        COALESCE(SAFE_CAST(CERRAR_CORTINA AS FLOAT64),0)+
        COALESCE(SAFE_CAST(PAPER_W        AS FLOAT64),0)
    )/60, 4) AS RECIBO_HRS,
    ROUND((
        COALESCE(SAFE_CAST(LLEGADA_A_TRAFICO    AS FLOAT64),0)+
        COALESCE(SAFE_CAST(DURACION_DE_SERVICIO AS FLOAT64),0)+
        COALESCE(SAFE_CAST(SALIDA_DE_CD         AS FLOAT64),0)
    )/60, 4) AS LOS_HRS
FROM `{BQ_TABLE}`
WHERE ARRIVAL_DATE BETWEEN '2026-07-01' AND '2026-07-31'
  AND SAFE_CAST(CEDIS AS INT64) = {CEDIS}
  AND UPPER(VENDOR) LIKE '%BONAFONT%'
  AND UPPER(TRIM(TIPO_CITA)) IN ('PROVEEDOR','USO CORRECTO')
  AND SAFE_CAST(SALIDA_DE_CD AS FLOAT64) > 0
ORDER BY SALIDA_HRS DESC
"""

print("Consultando BQ — Bonafont Chalco julio 2026 ...")
df = bigquery.Client().query(q).to_dataframe()
print(f"  {len(df):,} citas con salida registrada")

# ─── Helpers ──────────────────────────────────────────────────────────────────
def fmt_hrs(h) -> str:
    if h is None or h == "":  return ""
    try:   h = float(h)
    except: return ""
    if np.isnan(h) or h <= 0: return ""
    hh = int(h); mm = round((h - hh) * 60)
    if mm == 60: hh += 1; mm = 0
    return f"{hh}:{mm:02d}"

HRS_COLS = ["LLEGADA_HRS","ABRIR_CORTINA","CERRAR_CORTINA",
            "PAPER_W","SALIDA_HRS","SERVICIO_HRS","RECIBO_HRS","LOS_HRS"]

COLS = ["FECHA","CITA","VENDOR","CEDIS","NOMBRE_CEDIS","TIPO_CITA",
        "CITAS_CORRECTAS","SW","PUNTUALIDAD","DIFERENCIA_MIN",
        "LLEGADA_HRS","ABRIR_CORTINA","CERRAR_CORTINA","PAPER_W",
        "SALIDA_HRS","SERVICIO_HRS","RECIBO_HRS","LOS_HRS"]

COL_W = {
    "FECHA":13,"CITA":16,"VENDOR":30,"CEDIS":8,"NOMBRE_CEDIS":22,
    "TIPO_CITA":16,"CITAS_CORRECTAS":10,"SW":6,
    "PUNTUALIDAD":18,"DIFERENCIA_MIN":14,
    "LLEGADA_HRS":13,"ABRIR_CORTINA":13,"CERRAR_CORTINA":14,
    "PAPER_W":10,"SALIDA_HRS":13,"SERVICIO_HRS":14,
    "RECIBO_HRS":13,"LOS_HRS":12,
}

HDR_HEX = "1F497D"   # azul Chalco
ALT_HEX = "DCE6F1"
TOT_HEX = "9DC3E6"

thin = Side(style="thin", color="CCCCCC")
BRD  = Border(left=thin, right=thin, top=thin, bottom=thin)

_ANTICIP = {"1 DIA ANTES","12-24 HORAS ANTES","6-12 HORAS ANTES","1-6 HORAS ANTES"}

def punt_fill_font(v):
    v = str(v).upper().strip() if v else ""
    if v in _ANTICIP:
        return PatternFill("solid", fgColor="E2EFDA"), Font(color="375623", size=9)
    if v == "A TIEMPO":
        return PatternFill("solid", fgColor="FFEB9C"), Font(color="9C5700", size=9)
    if "DESPU" in v or "TARDE" in v:
        return PatternFill("solid", fgColor="FFC7CE"), Font(color="9C0006", size=9)
    return None, Font(size=9)

# umbral "tiempo alto": por encima del promedio de salida
avg_salida = df["SALIDA_HRS"].mean()
p75_salida = df["SALIDA_HRS"].quantile(0.75)
print(f"  Promedio SALIDA: {fmt_hrs(avg_salida)} | P75: {fmt_hrs(p75_salida)}")

# ─── Excel ────────────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "BONAFONT CHALCO JULIO"
ncols = len(COLS)

# Titulo
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
t = ws.cell(1, 1,
    f"BONAFONT - CHALCO (CEDIS {CEDIS}) | Julio 2026 | Ent. Directas + Uso Correcto | "
    f"{len(df):,} citas | ordenado por SALIDA desc | tiempos hh:mm")
t.fill = PatternFill("solid", fgColor=HDR_HEX)
t.font = Font(bold=True, color="FFFFFF", size=10)
t.alignment = Alignment(horizontal="center", vertical="center")

# Cabecera
HDR_LABELS = {c: c for c in COLS}
HDR_LABELS["SALIDA_HRS"] = "SALIDA_HRS\n(ordenado desc)"

for j, col in enumerate(COLS, 1):
    c = ws.cell(2, j, HDR_LABELS.get(col, col))
    c.fill = PatternFill("solid", fgColor=HDR_HEX)
    c.font = Font(bold=True, color="FFFFFF", size=9)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BRD

# Datos
_RED_FILL = PatternFill("solid", fgColor="FFE0E0")   # salida muy alta (> P75)
_ORG_FILL = PatternFill("solid", fgColor="FFF2CC")   # salida alta (> promedio)
_ALT_FILL = PatternFill("solid", fgColor=ALT_HEX)

for i, (_, row) in enumerate(df.iterrows()):
    ri = i + 3
    salida_val = row.get("SALIDA_HRS", 0) or 0

    if salida_val >= p75_salida:
        row_fill = _RED_FILL
    elif salida_val >= avg_salida:
        row_fill = _ORG_FILL
    else:
        row_fill = _ALT_FILL if i % 2 == 0 else None

    for j, col in enumerate(COLS, 1):
        val = row.get(col)
        if col in HRS_COLS:
            cell_val = fmt_hrs(val) or None
        else:
            cell_val = None if (val is None or (isinstance(val, float) and np.isnan(val))) else val

        dc = ws.cell(ri, j, cell_val)

        if col == "PUNTUALIDAD":
            pf, pfo = punt_fill_font(val)
            if pf: dc.fill = pf
            dc.font = pfo
        elif col == "SALIDA_HRS":
            dc.fill = _RED_FILL if salida_val >= p75_salida else (_ORG_FILL if salida_val >= avg_salida else (row_fill or PatternFill()))
            dc.font = Font(bold=True, size=9, color="C00000" if salida_val >= p75_salida else "7F6000")
        else:
            dc.font = Font(bold=(col in ("CITA","RECIBO_HRS","LOS_HRS")), size=9)
            if row_fill: dc.fill = row_fill

        dc.alignment = Alignment(horizontal="left" if col in ("FECHA","VENDOR","NOMBRE_CEDIS","PUNTUALIDAD") else "center",
                                  vertical="center")
        dc.border = BRD

# Fila de totales
row_tot = len(df) + 4
ws.row_dimensions[row_tot - 1].height = 5   # separador

tot_fill = PatternFill("solid", fgColor="2E4057")
tot_font = Font(bold=True, color="FFFFFF", size=9)
tot_lbl  = ws.cell(row_tot, 1, f"PROMEDIO ({len(df)} citas)")
tot_lbl.fill = tot_fill; tot_lbl.font = tot_font
tot_lbl.alignment = Alignment(horizontal="center", vertical="center")
tot_lbl.border = BRD

for j, col in enumerate(COLS, 1):
    if j == 1: continue
    if col in HRS_COLS:
        vals = df[col].dropna(); vals = vals[vals > 0]
        avg  = fmt_hrs(vals.mean()) if len(vals) else ""
        cell_val = avg or None
    else:
        cell_val = None
    dc = ws.cell(row_tot, j, cell_val)
    dc.fill = tot_fill; dc.font = tot_font
    dc.alignment = Alignment(horizontal="center", vertical="center")
    dc.border = BRD

# Leyenda de colores
ley_row = row_tot + 2
ws.cell(ley_row,   1, "ROJO:").font   = Font(bold=True, size=9)
ws.cell(ley_row,   2, f"Salida >= P75 ({fmt_hrs(p75_salida)})").font = Font(size=9)
ws.cell(ley_row,   1).fill = _RED_FILL
ws.cell(ley_row+1, 1, "AMARILLO:").font = Font(bold=True, size=9)
ws.cell(ley_row+1, 2, f"Salida >= Promedio ({fmt_hrs(avg_salida)})").font = Font(size=9)
ws.cell(ley_row+1, 1).fill = _ORG_FILL

# Anchos y freeze
for j, col in enumerate(COLS, 1):
    ws.column_dimensions[get_column_letter(j)].width = COL_W.get(col, 12)
ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 30
ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}2"

wb.save(OUT_FILE)
print(f"\nListo -> {OUT_FILE.name}")
print(f"  Citas con salida > P75 ({fmt_hrs(p75_salida)}): {int((df['SALIDA_HRS'] >= p75_salida).sum())}")
print(f"  Citas con salida > prom ({fmt_hrs(avg_salida)}): {int((df['SALIDA_HRS'] >= avg_salida).sum())}")
