"""
export_detalle_jul1_10.py
Detalle cita por cita Jul 1-10 2026 (Directo + BKHL) desde BigQuery.
"""
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.cloud import bigquery

BASE = Path(__file__).parent
BQ_TABLE = "wmt-intl-supplychain-gcp-prod.MR101_WM_AD_HOC.SCH_YMS_SEMANAL"
FECHA_INI = "2026-07-01"
FECHA_FIN = "2026-07-10"

VENDORS_KW = [
    "BONAFONT", "PEPSICO", "NIAGARA", "SUPREMA", "FRABEL",
    "HERDEZ", "JUGOS DEL VALLE", "KIMBERLY", "NESTLE", "MONDELEZ",
    "PROCTER", "SANTA CLARA", "UNILEVER",
]

VENDOR_DISPLAY = {
    "BONAFONT":        "BONAFONT SA CV",
    "PEPSICO":         "COMERC PEPSICO MEXICO S RL CV",
    "NIAGARA":         "EMBOTELLAD NIAGARA D MX S RLCV",
    "SUPREMA":         "ENVASADORA LA SUPREMA SA DE CV",
    "FRABEL":          "FRABEL SA DE CV",
    "HERDEZ":          "HERDEZ SA DE CV",
    "JUGOS DEL VALLE": "JUGOS DEL VALLE SAPI DE CV",
    "KIMBERLY":        "KIMBERLY CLARK MEXICO SA B CV",
    "NESTLE":          "MARCAS NESTLE SA CV",
    "MONDELEZ":        "MONDELEZ MEXICO S DE RL DE CV",
    "PROCTER":         "PROCTER AND GAMBLE MEXICO INC",
    "SANTA CLARA":     "SANTA CLARA MERC PACHU S RL CV",
    "UNILEVER":        "UNILEVER DE MEXICO S RL CV",
}

CD_GROUPS = {
    "Cuautitlan N1": [7494],
    "Cuautitlan N2": [7464],
    "Megapark SMO":  [6388],
    "Santa Barbara": [7457, 7482],
    "Chihuahua":     [4640, 5780],
    "Culiacan":      [4971, 7455, 7487],
    "Mexicali":      [4924, 6140],
    "Monterrey":     [4995, 7461, 7490, 8806],
    "Chalco":        [7459, 7471, 7505],
    "Guadalajara":   [5907, 6238, 7460, 7493],
    "Merida":        [4188, 7103, 7506],
    "Villahermosa":  [6550, 7453, 7468],
}
CEDIS_TO_CD = {c: cd for cd, cs in CD_GROUPS.items() for c in cs}


def label_vendor(v: str):
    v = str(v).upper()
    for kw in VENDORS_KW:
        if kw in v:
            return VENDOR_DISPLAY[kw]
    return None


def query_detalle(tipo_filter: str) -> pd.DataFrame:
    client = bigquery.Client()
    q = f"""
    SELECT
        APPOINTMENT_NBR,
        ARRIVAL_DATE,
        CEDIS,
        NOMBRE_CEDIS,
        LOCACION,
        VENDOR,
        TIPO_CITA,
        CITAS_CORRECTAS,
        SW,
        LLEGADA_A_TRAFICO,
        ABRIR_CORTINA,
        CERRAR_CORTINA,
        PAPER_W,
        SALIDA_DE_CD,
        DURACION_DE_SERVICIO,
        ROUND(
            (
              COALESCE(SAFE_CAST(LLEGADA_A_TRAFICO    AS FLOAT64), 0) +
              COALESCE(SAFE_CAST(DURACION_DE_SERVICIO AS FLOAT64), 0) +
              COALESCE(SAFE_CAST(SALIDA_DE_CD         AS FLOAT64), 0)
            ) / 60.0, 4
        ) AS LOS_HRS
    FROM `{BQ_TABLE}`
    WHERE ARRIVAL_DATE BETWEEN '{FECHA_INI}' AND '{FECHA_FIN}'
      AND {tipo_filter}
    ORDER BY ARRIVAL_DATE, CEDIS, VENDOR
    """
    return client.query(q).to_dataframe()


def enrich(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["CEDIS"] = pd.to_numeric(df["CEDIS"], errors="coerce")
    df["VENDOR_LABEL"] = df["VENDOR"].apply(label_vendor)
    df["CD_GRUPO"] = df["CEDIS"].map(CEDIS_TO_CD)
    front = [
        "ARRIVAL_DATE", "APPOINTMENT_NBR", "VENDOR_LABEL", "VENDOR",
        "CD_GRUPO", "CEDIS", "NOMBRE_CEDIS", "LOCACION",
        "TIPO_CITA", "CITAS_CORRECTAS", "SW",
        "LLEGADA_A_TRAFICO", "ABRIR_CORTINA", "CERRAR_CORTINA",
        "PAPER_W", "SALIDA_DE_CD", "DURACION_DE_SERVICIO", "LOS_HRS",
    ]
    return df[[c for c in front if c in df.columns]]


print(f"Consultando BigQuery: {FECHA_INI} al {FECHA_FIN}")

print("  Directo (Proveedor / Cita Nueva)...")
df_dir_raw = query_detalle("UPPER(TRIM(TIPO_CITA)) IN ('PROVEEDOR', 'CITA NUEVA')")
print(f"    {len(df_dir_raw):,} registros totales")

print("  BKHL (Backhaul)...")
df_bkhl_raw = query_detalle("UPPER(TRIM(TIPO_CITA)) IN ('BACKHAUL', 'CNV BACKHAUL', 'REPRO BKH')")
print(f"    {len(df_bkhl_raw):,} registros totales")

df_dir  = enrich(df_dir_raw)
df_bkhl = enrich(df_bkhl_raw)

df_dir_v  = df_dir[df_dir["VENDOR_LABEL"].notna()].copy()
df_bkhl_v = df_bkhl[df_bkhl["VENDOR_LABEL"].notna()].copy()

print(f"\n  Directo  vendors objetivo: {len(df_dir_v):,} citas")
print(f"  BKHL     vendors objetivo: {len(df_bkhl_v):,} citas")


# ─── Excel ───────────────────────────────────────────────────────────────────
COL_WIDTHS = {
    "ARRIVAL_DATE": 13, "APPOINTMENT_NBR": 16, "VENDOR_LABEL": 30,
    "VENDOR": 40, "CD_GRUPO": 16, "CEDIS": 8, "NOMBRE_CEDIS": 26,
    "LOCACION": 22, "TIPO_CITA": 15, "CITAS_CORRECTAS": 8, "SW": 6,
    "LLEGADA_A_TRAFICO": 13, "ABRIR_CORTINA": 13, "CERRAR_CORTINA": 13,
    "PAPER_W": 10, "SALIDA_DE_CD": 13, "DURACION_DE_SERVICIO": 16, "LOS_HRS": 10,
}


def write_sheet(ws, df: pd.DataFrame, titulo: str, hdr_hex: str, alt_hex: str):
    ws.title = titulo[:31]
    hdr_fill = PatternFill("solid", fgColor=hdr_hex)
    hdr_font = Font(bold=True, color="FFFFFF", size=9)
    alt_fill = PatternFill("solid", fgColor=alt_hex)
    norm_f   = Font(size=9)
    bold_f   = Font(bold=True, size=9)
    center   = Alignment(horizontal="center", vertical="center")
    left_a   = Alignment(horizontal="left",   vertical="center")
    thin     = Side(style="thin", color="CCCCCC")
    brd      = Border(left=thin, right=thin, top=thin, bottom=thin)

    cols = list(df.columns)
    ncols = len(cols)

    # Fila titulo
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(1, 1, f"DETALLE CITAS — {titulo} | {FECHA_INI} al {FECHA_FIN} | {len(df):,} registros")
    t.fill = hdr_fill
    t.font = Font(bold=True, color="FFFFFF", size=10)
    t.alignment = center

    # Cabecera columnas
    for j, col in enumerate(cols, 1):
        c = ws.cell(2, j, col)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = center
        c.border = brd

    # Datos
    LEFT_COLS = {"VENDOR_LABEL", "VENDOR", "NOMBRE_CEDIS", "LOCACION", "TIPO_CITA"}
    BOLD_COLS = {"VENDOR_LABEL", "LOS_HRS"}
    for i, row in enumerate(df.itertuples(index=False), 3):
        rf = alt_fill if (i % 2 == 0) else None
        for j, val in enumerate(row, 1):
            col = cols[j - 1]
            cell_val = None if (val is None or (isinstance(val, float) and pd.isna(val))) else val
            dc = ws.cell(i, j, cell_val)
            dc.font = bold_f if col in BOLD_COLS else norm_f
            dc.alignment = left_a if col in LEFT_COLS else center
            dc.border = brd
            if rf:
                dc.fill = rf

    # Anchos y alturas
    for j, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = COL_WIDTHS.get(col, 12)
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}2"


out = BASE / "detalle_citas_jul1_10.xlsx"
print(f"\nEscribiendo {out} ...")
wb = Workbook()

write_sheet(wb.active,         df_dir_v,  "DIRECTO",  "1F4E79", "D9E1F2")
write_sheet(wb.create_sheet(), df_bkhl_v, "BKHL",     "833C00", "FCE4D6")

wb.save(out)
print(f"Excel guardado: {out}")
print(f"  DIRECTO : {len(df_dir_v):,} citas")
print(f"  BKHL    : {len(df_bkhl_v):,} citas")
