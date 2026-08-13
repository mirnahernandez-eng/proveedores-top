"""
export_bkhl_todos.py
BKHL 2026 YTD (2026-01-01 → 2026-07-24, SW1–SW25)
Todos los CEDIS — 75 proveedores TOP.
Salida: bkhl_todos_proveedores_SW25.xlsx
  Hoja DETALLE : cita por cita ordenada por vendor / fecha
  Hoja RESUMEN : 1 fila por vendor con totales y promedios
"""
import re
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.cloud import bigquery

BASE      = Path(__file__).parent
BQ_TABLE  = "wmt-intl-supplychain-gcp-prod.MR101_WM_AD_HOC.SCH_YMS_SEMANAL"
FECHA_INI = "2026-01-01"
FECHA_FIN = "2026-07-24"
OUT_FILE  = BASE / "bkhl_todos_proveedores_SW25.xlsx"

# ─── (display_name, [python_re_patterns], bq_keyword) ────────────────────────
VENDOR_RULES = [
    ("4E GLOBAL SAPI DE CV",                              ["4E GLOBAL"],               "4E GLOBAL"),
    ("ABSORMEX CMPC TISSUE S.A. DE CV",                   ["ABSORMEX", "CMPC TISSUE"], "ABSORMEX"),
    ("ACH FOODS MEXICO S DE RL DE CV",                    ["ACH FOODS"],               "ACH FOODS"),
    ("AJEMEX SA DE CV",                                   ["AJEMEX"],                  "AJEMEX"),
    ("AKSI HERRAMIENTAS SA DE CV",                        ["AKSI"],                    "AKSI"),
    ("BIO PAPPEL SCRIBE SA DE CV",                        ["SCRIBE", "BIO PAPPEL"],    "SCRIBE"),
    ("BONAFONT SA DE CV",                                 ["BONAFONT"],                "BONAFONT"),
    ("BRONCOLIN SA DE CV",                                ["BRONCOLIN"],               "BRONCOLIN"),
    ("CESARFER SA DE CV",                                 ["CESARFER"],                "CESARFER"),
    ("CH & ML ELECTRIC MEXICO S DE RL DE CV",             ["CH.{0,4}ML.{0,4}ELEC"],   "CH.*ML.*ELEC"),
    ("CHURCH & DWIGHT S DE RL DE CV",                     ["CHURCH.{0,8}DWIGHT"],      "CHURCH.*DWIGHT"),
    ("CIA COMERCIAL HERDEZ SA DE CV",                     ["HERDEZ"],                  "HERDEZ"),
    ("CIA INTERNAC COMERCIO SAPI CV",                     ["INTERNAC.{0,12}COMERC"],   "INTERNAC.*COMERC"),
    ("COLCHONES WENDY SA DE CV",                          ["WENDY"],                   "WENDY"),
    ("COMERC PEPSICO MEXICO S RL CV",                     ["PEPSICO"],                 "PEPSICO"),
    ("COMERCIAL 100 MEXICAN SA DE CV",                    ["100 MEXICAN"],             "100 MEXICAN"),
    ("COMERCIALIZADORA ELORO SA",                         ["ELORO"],                   "ELORO"),
    ("COMPAÑIA COMERCIALIZADORA PRODIN CENTRO SA DE CV",  ["PRODIN"],                  "PRODIN"),
    ("CONAGRA FOODS MEXICO SA DE CV",                     ["CONAGRA"],                 "CONAGRA"),
    ("CONSERVAS LA COSTENA SA DE CV",                     ["COSTE.A", "COSTENA"],      "COSTENA"),
    ("CORPORACION GAIRET SA DE CV",                       ["GAIRET"],                  "GAIRET"),
    ("CRISA LIBBEY COMERCIAL S RL CV",                    ["CRISA", "LIBBEY"],         "CRISA"),
    ("CUETARA DISTRIBUCION SA DE CV",                     ["CUETARA"],                 "CUETARA"),
    ("DASAVENA GOURMET SA DE CV",                         ["DASAVENA"],                "DASAVENA"),
    ("EFFEM MEXICO INC Y CIA S NC CV",                    ["EFFEM"],                   "EFFEM"),
    ("EMBOTELLADORA NIAGARA SA DE CV",                    ["NIAGARA"],                 "NIAGARA"),
    ("ESPEJOS INTELIGENTES SA DE CV",                     ["ESPEJOS"],                 "ESPEJOS"),
    ("FAB DE JABON LA CORONA SA DE CV",                   ["LA CORONA", "JABON.{0,10}CORONA"], "LA CORONA"),
    ("FACTOR PESCA SA DE CV",                             ["FACTOR PESCA"],            "FACTOR PESCA"),
    ("FANTASY RUZ S.A. DE C.V.",                          ["FANTASY RUZ"],             "FANTASY RUZ"),
    ("FRABEL SA DE CV",                                   ["FRABEL"],                  "FRABEL"),
    ("GANAD PROD DE LECHE PURA SA DE CV",                 ["LECHE PURA"],              "LECHE PURA"),
    ("GRUPO TAIFELDS SA DE CV",                           ["TAIFELDS"],                "TAIFELDS"),
    ("HALEON CONSUMER S DE RL DE CV",                     ["HALEON"],                  "HALEON"),
    ("HAPPY FLOWER MEXICANA SA DE CV",                    ["HAPPY FLOWER"],            "HAPPY FLOWER"),
    ("HENKEL CAPITAL SA DE CV",                           ["HENKEL"],                  "HENKEL"),
    ("HFC PRESTIGE INTERNATIONAL S D",                    ["HFC PRESTIGE"],            "HFC PRESTIGE"),
    ("HISENSE MEXICO S DE RL DE CV",                      ["HISENSE"],                 "HISENSE"),
    ("IMPERCAUCHO SA DE CV",                              ["IMPERCAUCHO"],             "IMPERCAUCHO"),
    ("IMPULSORA CAMPIRANO SA DE CV",                      ["CAMPIRANO"],               "CAMPIRANO"),
    ("IND NAC DE DETERGENTES SA DE CV",                   ["NAC.{0,10}DETERG"],        "DETERGENTES"),
    ("INDUSTRIAS OVARB SA DE CV",                         ["OVARB"],                   "OVARB"),
    ("INDUSTRIAS SALCOM SA DE CV",                        ["SALCOM"],                  "SALCOM"),
    ("JESSY INTERNACIONAL SA DE CV",                      ["JESSY"],                   "JESSY"),
    ("JUVASA SERVICIOS SA DE CV",                         ["JUVASA"],                  "JUVASA"),
    ("KIMBERLY CLARK DE MEX SA DE CV",                    ["KIMBERLY"],                "KIMBERLY"),
    ("KSMV CAPITAL SAPI DE",                              ["KSMV"],                    "KSMV"),
    ("LA MASCOTA SA DE CV",                               ["LA MASCOTA"],              "LA MASCOTA"),
    ("MARCAS NESTLE SA DE CV",                            ["NESTLE"],                  "NESTLE"),
    ("MATTEL DE MEXICO SA DE CV",                         ["MATTEL"],                  "MATTEL"),
    ("MONDELEZ MEXICO S DE RL DE CV",                     ["MONDELEZ"],                "MONDELEZ"),
    ("NEWELL BRANDS DE MEXICO SA DE C",                   ["NEWELL"],                  "NEWELL"),
    ("PEÑAFIEL BEBIDAS SA DE CV",                         ["PE.AFIEL", "PENAFIEL"],    "PENAFIEL"),
    ("PLAYERAS SOURCE SA DE CV",                          ["PLAYERAS SOURCE"],         "PLAYERAS SOURCE"),
    ("POLYCHEM SA DE CV",                                 ["POLYCHEM"],                "POLYCHEM"),
    ("PROBEMEX SA DE CV",                                 ["PROBEMEX"],                "PROBEMEX"),
    ("PROCTER & GAMBLE MEXICO S DE RL DE CV",             ["PROCTER"],                 "PROCTER"),
    ("PROD ALIMENT LA MODERNA SA DE CV",                  ["LA MODERNA"],              "LA MODERNA"),
    ("PRODUCTOS INTERNACIONALES MABE SA DE CV",           ["MABE"],                    "MABE"),
    ("PROXIMO NATAL SAPI DE CV",                          ["PROXIMO"],                 "PROXIMO"),
    ("QUALAMEX SA DE CV",                                 ["QUALAMEX"],                "QUALAMEX"),
    ("RAMIREZ ZUNIGA LAURA",                              ["RAMIREZ ZUNIGA"],          "RAMIREZ ZUNIGA"),
    ("RECKITT BENCKISER MEXICO",                          ["RECKITT"],                 "RECKITT"),
    ("REGALOS SIGLO XXI SA DE CV",                        ["REGALOS SIGLO"],           "REGALOS SIGLO"),
    ("SANTA CLARA MERCANTIL DE PACHUCA S DE RL DE CV",    ["SANTA CLARA"],             "SANTA CLARA"),
    ("SANTUL HERRAMIENTAS SA DE CV",                      ["SANTUL"],                  "SANTUL"),
    ("SCHETTINO HNOS SRL DE CV",                          ["SCHETTINO"],               "SCHETTINO"),
    ("SERVICIOS NUTRICIONALES MEAD JOHNSON S DE RL DE CV",["MEAD JOHNSON"],            "MEAD JOHNSON"),
    ("SUPER FOODS FACTORY",                               ["SUPER FOODS"],             "SUPER FOODS"),
    ("TECNISPICE SA DE CV",                               ["TECNISPICE"],              "TECNISPICE"),
    ("TRESMONTES LUCCHETTI MEX SA DE CV",                 ["TRESMONTES", "LUCCHETTI"], "TRESMONTES"),
    ("UNILEVER DE MEXICO S RL CV",                        ["UNILEVER"],                "UNILEVER"),
    ("VASCONIA BRANDS SA DE CV",                          ["VASCONIA"],                "VASCONIA"),
    ("VCT & DG MEXICO SA DE CV",                          ["VCT.{0,5}DG"],             "VCT"),
    ("VIDRIERA SANTOS SA DE CV",                          ["VIDRIERA SANTOS"],         "VIDRIERA SANTOS"),
]

VENDOR_NAMES = [r[0] for r in VENDOR_RULES]
_BQ_PAT      = "|".join("(?:" + r[2] + ")" for r in VENDOR_RULES)

NUM_COLS = [
    "LLEGADA_A_TRAFICO", "ABRIR_CORTINA", "CERRAR_CORTINA",
    "PAPER_W", "SALIDA_DE_CD", "DURACION_DE_SERVICIO",
    "RECIBO_HRS", "LOS_HRS",
]
COLS_DET = [
    "FECHA", "# CITA", "VENDOR", "CEDIS", "CEDIS_NOMBRE",
    "TIPO_CITA", "SW",
    "LLEGADA_A_TRAFICO", "ABRIR_CORTINA", "CERRAR_CORTINA",
    "PAPER_W", "SALIDA_DE_CD", "DURACION_DE_SERVICIO",
    "RECIBO_HRS", "LOS_HRS",
]
COL_W = {
    "FECHA": 13, "# CITA": 16, "VENDOR": 40, "CEDIS": 8,
    "CEDIS_NOMBRE": 22, "TIPO_CITA": 16, "SW": 6,
    "LLEGADA_A_TRAFICO": 14, "ABRIR_CORTINA": 13, "CERRAR_CORTINA": 14,
    "PAPER_W": 10, "SALIDA_DE_CD": 13, "DURACION_DE_SERVICIO": 18,
    "RECIBO_HRS": 12, "LOS_HRS": 12,
}


# ─── BigQuery ──────────────────────────────────────────────────────────────────
def query_bkhl() -> pd.DataFrame:
    client = bigquery.Client()
    q = (
        "SELECT\n"
        "    ARRIVAL_DATE                                   AS FECHA,\n"
        "    APPOINTMENT_NBR                                AS `# CITA`,\n"
        "    VENDOR,\n"
        "    SAFE_CAST(CEDIS AS INT64)                      AS CEDIS,\n"
        "    NOMBRE_CEDIS                                   AS CEDIS_NOMBRE,\n"
        "    TIPO_CITA,\n"
        "    SW,\n"
        "    ROUND(SAFE_CAST(LLEGADA_A_TRAFICO    AS FLOAT64)/60,4) AS LLEGADA_A_TRAFICO,\n"
        "    ROUND(SAFE_CAST(ABRIR_CORTINA        AS FLOAT64)/60,4) AS ABRIR_CORTINA,\n"
        "    ROUND(SAFE_CAST(CERRAR_CORTINA       AS FLOAT64)/60,4) AS CERRAR_CORTINA,\n"
        "    ROUND(SAFE_CAST(PAPER_W              AS FLOAT64)/60,4) AS PAPER_W,\n"
        "    ROUND(SAFE_CAST(SALIDA_DE_CD         AS FLOAT64)/60,4) AS SALIDA_DE_CD,\n"
        "    ROUND(SAFE_CAST(DURACION_DE_SERVICIO AS FLOAT64)/60,4) AS DURACION_DE_SERVICIO,\n"
        "    ROUND((\n"
        "        COALESCE(SAFE_CAST(ABRIR_CORTINA  AS FLOAT64),0)+\n"
        "        COALESCE(SAFE_CAST(CERRAR_CORTINA AS FLOAT64),0)+\n"
        "        COALESCE(SAFE_CAST(PAPER_W        AS FLOAT64),0)\n"
        "    )/60,4) AS RECIBO_HRS,\n"
        "    ROUND((\n"
        "        COALESCE(SAFE_CAST(LLEGADA_A_TRAFICO    AS FLOAT64),0)+\n"
        "        COALESCE(SAFE_CAST(DURACION_DE_SERVICIO AS FLOAT64),0)+\n"
        "        COALESCE(SAFE_CAST(SALIDA_DE_CD         AS FLOAT64),0)\n"
        "    )/60,4) AS LOS_HRS\n"
        "FROM `" + BQ_TABLE + "`\n"
        "WHERE ARRIVAL_DATE BETWEEN '" + FECHA_INI + "' AND '" + FECHA_FIN + "'\n"
        "  AND UPPER(TRIM(TIPO_CITA)) IN ('BACKHAUL','CNV BACKHAUL','REPRO BKH')\n"
        "  AND REGEXP_CONTAINS(UPPER(VENDOR), r'" + _BQ_PAT + "')\n"
        "  AND (\n"
        "        COALESCE(SAFE_CAST(ABRIR_CORTINA        AS FLOAT64),0)+\n"
        "        COALESCE(SAFE_CAST(CERRAR_CORTINA       AS FLOAT64),0)+\n"
        "        COALESCE(SAFE_CAST(PAPER_W              AS FLOAT64),0)+\n"
        "        COALESCE(SAFE_CAST(LLEGADA_A_TRAFICO    AS FLOAT64),0)+\n"
        "        COALESCE(SAFE_CAST(DURACION_DE_SERVICIO AS FLOAT64),0)+\n"
        "        COALESCE(SAFE_CAST(SALIDA_DE_CD         AS FLOAT64),0)\n"
        "      ) > 0\n"
        "ORDER BY ARRIVAL_DATE, VENDOR\n"
    )
    return client.query(q).to_dataframe()


# ─── Matching & enrich ────────────────────────────────────────────────────────
def match_vendor(raw: str) -> str | None:
    raw_up = str(raw).upper()
    for display, patterns, _ in VENDOR_RULES:
        for pat in patterns:
            if re.search(pat, raw_up):
                return display
    return None


def enrich(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["VENDOR"] = df["VENDOR"].apply(match_vendor)
    df = df[df["VENDOR"].notna()].copy()
    df["VENDOR"] = pd.Categorical(df["VENDOR"], categories=VENDOR_NAMES, ordered=True)
    return df.sort_values(["VENDOR", "FECHA", "CEDIS"]).reset_index(drop=True)


# ─── Totales ──────────────────────────────────────────────────────────────────
def make_totals(df: pd.DataFrame) -> pd.DataFrame:
    rows, grand_avgs = [], {c: [] for c in NUM_COLS}
    for vendor in VENDOR_NAMES:
        sub = df[df["VENDOR"] == vendor]
        if sub.empty:
            continue
        base = sub[sub["RECIBO_HRS"].notna() & (sub["RECIBO_HRS"] > 0)]
        if base.empty:
            continue
        row = {"FECHA": f"TOTAL — {vendor}", "# CITA": len(sub),
               "VENDOR": vendor, "CEDIS": "", "CEDIS_NOMBRE": "",
               "TIPO_CITA": "", "SW": ""}
        for col in ("ABRIR_CORTINA", "CERRAR_CORTINA", "PAPER_W", "DURACION_DE_SERVICIO"):
            avg = round(base[col].fillna(0).mean(), 4)
            row[col] = avg
            grand_avgs[col].append(avg)
        row["RECIBO_HRS"] = round(row["ABRIR_CORTINA"] + row["CERRAR_CORTINA"] + row["PAPER_W"], 4)
        grand_avgs["RECIBO_HRS"].append(row["RECIBO_HRS"])
        for col in ("LLEGADA_A_TRAFICO", "SALIDA_DE_CD", "LOS_HRS"):
            vals = sub[col].dropna()
            vals = vals[vals > 0]
            avg = round(vals.mean(), 4) if len(vals) else None
            row[col] = avg
            if avg is not None:
                grand_avgs[col].append(avg)
        rows.append(row)

    grand = {"FECHA": "TOTAL GENERAL", "# CITA": len(df),
             "VENDOR": "", "CEDIS": "", "CEDIS_NOMBRE": "", "TIPO_CITA": "", "SW": ""}
    for col in NUM_COLS:
        avgs = grand_avgs[col]
        grand[col] = round(sum(avgs) / len(avgs), 4) if avgs else None
    if grand.get("ABRIR_CORTINA") is not None:
        grand["RECIBO_HRS"] = round(
            grand["ABRIR_CORTINA"] + grand["CERRAR_CORTINA"] + grand["PAPER_W"], 4)
    rows.append(grand)
    return pd.DataFrame(rows, columns=COLS_DET)


def make_resumen(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for vendor in VENDOR_NAMES:
        sub = df[df["VENDOR"] == vendor]
        if sub.empty:
            continue
        base = sub[sub["RECIBO_HRS"].notna() & (sub["RECIBO_HRS"] > 0)]
        row = {"VENDOR": vendor, "# CITAS": len(sub)}
        for col in ("ABRIR_CORTINA", "CERRAR_CORTINA", "PAPER_W", "DURACION_DE_SERVICIO"):
            row[col] = round(base[col].fillna(0).mean(), 4) if not base.empty else None
        row["RECIBO_HRS"] = (round(row["ABRIR_CORTINA"] + row["CERRAR_CORTINA"] + row["PAPER_W"], 4)
                             if row.get("ABRIR_CORTINA") is not None else None)
        for col in ("LLEGADA_A_TRAFICO", "SALIDA_DE_CD", "LOS_HRS"):
            vals = sub[col].dropna()
            vals = vals[vals > 0]
            row[col] = round(vals.mean(), 4) if len(vals) else None
        rows.append(row)
    return pd.DataFrame(rows)


# ─── Helpers de formato ───────────────────────────────────────────────────────
def fmt_hrs(h) -> str:
    if h is None or (isinstance(h, float) and (np.isnan(h) or h <= 0)):
        return ""
    hh = int(h)
    mm = round((h - hh) * 60)
    if mm == 60:
        hh += 1; mm = 0
    return f"{hh}:{mm:02d}"


HDR_BLUE  = "0071CE"
HDR_DARK  = "004C97"
ALT_BLUE  = "E8F4FD"
TOT_FILL  = "FFC220"
GRAND_FILL= "004C97"


def _s() -> dict:
    thin = Side(style="thin", color="CCCCCC")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    return {
        "hdr":   PatternFill("solid", fgColor=HDR_BLUE),
        "alt":   PatternFill("solid", fgColor=ALT_BLUE),
        "tot":   PatternFill("solid", fgColor=TOT_FILL),
        "grand": PatternFill("solid", fgColor=GRAND_FILL),
        "wf":    Font(bold=True, color="FFFFFF", size=10),
        "hf":    Font(bold=True, color="FFFFFF", size=9),
        "bf":    Font(bold=True, size=9),
        "nf":    Font(size=9),
        "gf":    Font(bold=True, color="FFFFFF", size=9),
        "ctr":   Alignment(horizontal="center", vertical="center"),
        "lft":   Alignment(horizontal="left",   vertical="center"),
        "brd":   brd,
    }


# ─── DETALLE sheet ────────────────────────────────────────────────────────────
def write_detalle(ws, df: pd.DataFrame, df_tot: pd.DataFrame):
    ws.title = "DETALLE"
    s = _s()
    ncols = len(COLS_DET)
    LEFT = {"FECHA", "VENDOR", "CEDIS_NOMBRE", "TIPO_CITA"}

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(1, 1,
        f"BKHL TODOS PROVEEDORES | {FECHA_INI} al {FECHA_FIN} | "
        f"{len(df):,} registros | tiempos en HORAS (hh:mm)")
    t.fill = s["hdr"]; t.font = s["wf"]; t.alignment = s["ctr"]

    for j, col in enumerate(COLS_DET, 1):
        c = ws.cell(2, j, col)
        c.fill = s["hdr"]; c.font = s["hf"]
        c.alignment = s["ctr"]; c.border = s["brd"]

    row_i = 3
    prev_vendor = None
    for _, row in df.iterrows():
        vendor = row["VENDOR"]
        fill = s["alt"] if vendor != prev_vendor else None
        prev_vendor = vendor
        for j, col in enumerate(COLS_DET, 1):
            val = row.get(col)
            if col in NUM_COLS:
                cell_val = fmt_hrs(val) or None
            else:
                cell_val = None if (val is None or (isinstance(val, float) and np.isnan(val))) else val
            dc = ws.cell(row_i, j, cell_val)
            dc.font = s["bf"] if col in ("VENDOR", "RECIBO_HRS", "LOS_HRS") else s["nf"]
            dc.alignment = s["lft"] if col in LEFT else s["ctr"]
            dc.border = s["brd"]
            if fill:
                dc.fill = fill
        row_i += 1

    ws.row_dimensions[row_i].height = 5
    row_i += 1
    for _, row in df_tot.iterrows():
        is_grand = str(row.get("FECHA", "")).startswith("TOTAL GENERAL")
        fill = s["grand"] if is_grand else s["tot"]
        font = s["gf"] if is_grand else s["bf"]
        for j, col in enumerate(COLS_DET, 1):
            val = row.get(col)
            if col in NUM_COLS:
                cell_val = fmt_hrs(val) or None
            else:
                cell_val = None if (val is None or (isinstance(val, float) and np.isnan(val))) else val
            dc = ws.cell(row_i, j, cell_val)
            dc.fill = fill; dc.font = font
            dc.alignment = s["lft"] if col in LEFT else s["ctr"]
            dc.border = s["brd"]
        row_i += 1

    for j, col in enumerate(COLS_DET, 1):
        ws.column_dimensions[get_column_letter(j)].width = COL_W.get(col, 12)
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 26
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}2"


# ─── RESUMEN sheet ────────────────────────────────────────────────────────────
_RES_COLS = ["VENDOR", "# CITAS", "LOS_HRS", "RECIBO_HRS",
             "LLEGADA_A_TRAFICO", "DURACION_DE_SERVICIO", "SALIDA_DE_CD",
             "ABRIR_CORTINA", "CERRAR_CORTINA", "PAPER_W"]
_RES_W    = {"VENDOR": 44, "# CITAS": 10, "LOS_HRS": 12, "RECIBO_HRS": 12,
             "LLEGADA_A_TRAFICO": 14, "DURACION_DE_SERVICIO": 18,
             "SALIDA_DE_CD": 13, "ABRIR_CORTINA": 13,
             "CERRAR_CORTINA": 14, "PAPER_W": 10}
_RES_NUM  = {"LOS_HRS", "RECIBO_HRS", "LLEGADA_A_TRAFICO",
             "DURACION_DE_SERVICIO", "SALIDA_DE_CD",
             "ABRIR_CORTINA", "CERRAR_CORTINA", "PAPER_W"}


def write_resumen(ws, df_res: pd.DataFrame):
    ws.title = "RESUMEN"
    s = _s()
    ncols = len(_RES_COLS)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(1, 1,
        f"RESUMEN BKHL YTD 2026 (SW1–SW25) | {len(df_res)} proveedores | promedios en hh:mm")
    t.fill = s["hdr"]; t.font = s["wf"]; t.alignment = s["ctr"]

    headers = {
        "VENDOR": "PROVEEDOR", "# CITAS": "# CITAS",
        "LOS_HRS": "LOS PROM", "RECIBO_HRS": "RECIBO PROM",
        "LLEGADA_A_TRAFICO": "LLEGADA PROM", "DURACION_DE_SERVICIO": "DURACIÓN PROM",
        "SALIDA_DE_CD": "SALIDA PROM", "ABRIR_CORTINA": "ABRIR CORTINA",
        "CERRAR_CORTINA": "CERRAR CORTINA", "PAPER_W": "PAPER W",
    }
    for j, col in enumerate(_RES_COLS, 1):
        c = ws.cell(2, j, headers[col])
        c.fill = s["hdr"]; c.font = s["hf"]
        c.alignment = s["ctr"]; c.border = s["brd"]

    for i, (_, row) in enumerate(df_res.iterrows(), 3):
        fill = s["alt"] if i % 2 == 1 else None
        for j, col in enumerate(_RES_COLS, 1):
            val = row.get(col)
            if col in _RES_NUM:
                cell_val = fmt_hrs(val) or None
            else:
                cell_val = val
            dc = ws.cell(i, j, cell_val)
            dc.font = s["bf"] if col in ("VENDOR", "LOS_HRS") else s["nf"]
            dc.alignment = s["lft"] if col == "VENDOR" else s["ctr"]
            dc.border = s["brd"]
            if fill:
                dc.fill = fill

    for j, col in enumerate(_RES_COLS, 1):
        ws.column_dimensions[get_column_letter(j)].width = _RES_W.get(col, 12)
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 26
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}2"


# ─── Main ──────────────────────────────────────────────────────────────────────
print(f"Consultando BigQuery BKHL — {FECHA_INI} al {FECHA_FIN} ...")
df_raw = query_bkhl()
print(f"  {len(df_raw):,} registros raw de BQ")

df = enrich(df_raw)
print(f"  {len(df):,} registros con vendors objetivo")

found = df["VENDOR"].nunique()
print(f"  {found} de {len(VENDOR_NAMES)} vendors encontrados")

df_tot = make_totals(df)
df_res = make_resumen(df)

print(f"\nEscribiendo {OUT_FILE} ...")
wb = Workbook()
write_detalle(wb.active, df, df_tot)
write_resumen(wb.create_sheet(), df_res)
wb.save(OUT_FILE)

print(f"Listo: {OUT_FILE}")
print(f"  DETALLE : {len(df):,} citas")
print(f"  RESUMEN : {len(df_res)} proveedores")
