"""
export_uso_incorrecto_gdl_julio.py
Citas con TIPO_CITA incorrecto (SCAC INCORRECTO + CITA NUEVA)
Guadalajara - 6 CEDIS - Julio 2026
Identifica la Semana Walmart (SW)
"""
import warnings; warnings.filterwarnings("ignore")
import sys, numpy as np, pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.cloud import bigquery

sys.path.insert(0, str(Path(__file__).parent))

BASE     = Path(__file__).parent
BQ_TABLE = "wmt-intl-supplychain-gcp-prod.MR101_WM_AD_HOC.SCH_YMS_SEMANAL"
OUT_FILE = BASE / "uso_incorrecto_gdl_julio2026.xlsx"

CEDIS_GDL = {
    5907: "SSTK",
    6238: "SECOS SAM'S",
    6239: "PERECEDEROS SAM'S",
    7460: "SECOS BAE",
    7493: "SECOS AUTOSERVICIOS",
    7495: "PERECEDEROS AUTOSERVICIOS",
}
TIPOS_INCORRECTOS = "'SCAC INCORRECTO','CITA NUEVA'"

# ─── SW calendar (Sat→Fri, SW1 = primera semana del año fiscal Walmart) ────────
def fecha_a_sw(fecha: pd.Timestamp) -> int:
    """Calcula SW Walmart (sábado a viernes). SW1 empieza el sábado más cercano al 1-Feb."""
    # Walmart fiscal year starts ~Feb 1; pero en MX se usa año calendario.
    # Usamos la convencion del tablero: SW basada en semana del año (lunes=inicio)
    return int(fecha.strftime("%W"))   # semana ISO lunes-inicio, 00-53


# Mejor: usar la SW tal como viene de BQ (columna SW)
# ─── Query ────────────────────────────────────────────────────────────────────
cedis_s = ",".join(str(c) for c in CEDIS_GDL)
q = f"""
SELECT
    ARRIVAL_DATE                                             AS FECHA,
    APPOINTMENT_NBR                                          AS CITA,
    VENDOR,
    SAFE_CAST(CEDIS AS INT64)                                AS CEDIS,
    NOMBRE_CEDIS,
    UPPER(TRIM(TIPO_CITA))                                   AS TIPO_CITA,
    CITAS_CORRECTAS,
    SW,
    COALESCE(CITA_VS_LLEGADA,'SIN DATO')                    AS PUNTUALIDAD,
    ROUND(SAFE_CAST(DIFERENCIA       AS FLOAT64),2)         AS DIFERENCIA_MIN,
    ROUND(SAFE_CAST(LLEGADA_A_TRAFICO    AS FLOAT64)/60,4)  AS LLEGADA_HRS,
    ROUND(SAFE_CAST(ABRIR_CORTINA        AS FLOAT64)/60,4)  AS ABRIR_CORTINA,
    ROUND(SAFE_CAST(CERRAR_CORTINA       AS FLOAT64)/60,4)  AS CERRAR_CORTINA,
    ROUND(SAFE_CAST(PAPER_W              AS FLOAT64)/60,4)  AS PAPER_W,
    ROUND(SAFE_CAST(SALIDA_DE_CD         AS FLOAT64)/60,4)  AS SALIDA_HRS,
    ROUND(SAFE_CAST(DURACION_DE_SERVICIO AS FLOAT64)/60,4)  AS SERVICIO_HRS,
    ROUND((
        COALESCE(SAFE_CAST(ABRIR_CORTINA  AS FLOAT64),0)+
        COALESCE(SAFE_CAST(CERRAR_CORTINA AS FLOAT64),0)+
        COALESCE(SAFE_CAST(PAPER_W        AS FLOAT64),0)
    )/60,4)  AS RECIBO_HRS,
    ROUND((
        COALESCE(SAFE_CAST(LLEGADA_A_TRAFICO    AS FLOAT64),0)+
        COALESCE(SAFE_CAST(DURACION_DE_SERVICIO AS FLOAT64),0)+
        COALESCE(SAFE_CAST(SALIDA_DE_CD         AS FLOAT64),0)
    )/60,4)  AS LOS_HRS
FROM `{BQ_TABLE}`
WHERE ARRIVAL_DATE BETWEEN '2026-07-01' AND '2026-07-31'
  AND SAFE_CAST(CEDIS AS INT64) IN ({cedis_s})
  AND UPPER(TRIM(TIPO_CITA)) IN ({TIPOS_INCORRECTOS})
ORDER BY CEDIS, ARRIVAL_DATE, VENDOR
"""

print("Consultando BQ — uso incorrecto GDL julio 2026 ...")
df = bigquery.Client().query(q).to_dataframe()
print(f"  {len(df):,} citas con tipo incorrecto")

# Resumen por CEDIS y tipo
print(df.groupby(["CEDIS","NOMBRE_CEDIS","TIPO_CITA"]).size().reset_index(name="N").to_string(index=False))

# ─── Helpers ──────────────────────────────────────────────────────────────────
HRS_COLS  = ["LLEGADA_HRS","ABRIR_CORTINA","CERRAR_CORTINA",
             "PAPER_W","SALIDA_HRS","SERVICIO_HRS","RECIBO_HRS","LOS_HRS"]
COLS      = ["FECHA","SW","CITA","VENDOR","CEDIS","NOMBRE_CEDIS",
             "TIPO_CITA","CITAS_CORRECTAS","PUNTUALIDAD","DIFERENCIA_MIN",
             "LLEGADA_HRS","ABRIR_CORTINA","CERRAR_CORTINA","PAPER_W",
             "SALIDA_HRS","SERVICIO_HRS","RECIBO_HRS","LOS_HRS"]
COL_W     = {
    "FECHA":13,"SW":6,"CITA":16,"VENDOR":34,"CEDIS":8,"NOMBRE_CEDIS":24,
    "TIPO_CITA":18,"CITAS_CORRECTAS":10,"PUNTUALIDAD":18,"DIFERENCIA_MIN":14,
    "LLEGADA_HRS":13,"ABRIR_CORTINA":13,"CERRAR_CORTINA":14,
    "PAPER_W":10,"SALIDA_HRS":13,"SERVICIO_HRS":14,
    "RECIBO_HRS":13,"LOS_HRS":12,
}
LEFT_COLS = {"FECHA","VENDOR","NOMBRE_CEDIS","TIPO_CITA","PUNTUALIDAD"}
thin      = Side(style="thin", color="CCCCCC")
BRD       = Border(left=thin, right=thin, top=thin, bottom=thin)
HDR_HEX   = "833C00"

# Color por CEDIS
CEDIS_FILLS = {
    5907: PatternFill("solid", fgColor="FFF2CC"),  # amarillo
    6238: PatternFill("solid", fgColor="DCE6F1"),  # azul claro
    6239: PatternFill("solid", fgColor="E2EFDA"),  # verde claro
    7460: PatternFill("solid", fgColor="FCE4D6"),  # naranja claro
    7493: PatternFill("solid", fgColor="F5E6F5"),  # lila
    7495: PatternFill("solid", fgColor="DDEBF7"),  # azul pálido
}
TIPO_SCAC_FILL = PatternFill("solid", fgColor="FFC7CE")   # rojo — SCAC INCORRECTO
TIPO_SCAC_FONT = Font(color="9C0006", bold=True, size=9)
TIPO_NUEV_FILL = PatternFill("solid", fgColor="FFEB9C")   # amarillo — CITA NUEVA
TIPO_NUEV_FONT = Font(color="9C5700", bold=True, size=9)


def fmt_hrs(h) -> str:
    if h is None or h == "": return ""
    try:   h = float(h)
    except: return ""
    if np.isnan(h) or h <= 0: return ""
    hh = int(h); mm = round((h - hh) * 60)
    if mm == 60: hh += 1; mm = 0
    return f"{hh}:{mm:02d}"


# ─── Excel ────────────────────────────────────────────────────────────────────
wb = Workbook(); ws = wb.active
ws.title = "USO INCORRECTO GDL JUL"
ncols    = len(COLS)
hdr_fill = PatternFill("solid", fgColor=HDR_HEX)
tot_fill = PatternFill("solid", fgColor="2E4057")

# Titulo
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
t = ws.cell(1, 1,
    f"USO INCORRECTO (SCAC INCORRECTO + CITA NUEVA) | GUADALAJARA | "
    f"Julio 2026 | {len(df):,} citas | ordenado por CEDIS / Fecha")
t.fill = hdr_fill; t.font = Font(bold=True, color="FFFFFF", size=10)
t.alignment = Alignment(horizontal="center", vertical="center")

# Cabecera
for j, col in enumerate(COLS, 1):
    c = ws.cell(2, j, col)
    c.fill = hdr_fill; c.font = Font(bold=True, color="FFFFFF", size=9)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BRD

# Datos
prev_cedis = None
for i, (_, row) in enumerate(df.iterrows()):
    ri       = i + 3
    cedis_v  = int(row.get("CEDIS") or 0)
    tipo_v   = str(row.get("TIPO_CITA") or "").upper()
    row_fill = CEDIS_FILLS.get(cedis_v)

    # Separador visual entre CEDIS
    if cedis_v != prev_cedis and prev_cedis is not None:
        ws.row_dimensions[ri - 1].height = 6
    prev_cedis = cedis_v

    for j, col in enumerate(COLS, 1):
        val      = row.get(col)
        cell_val = (fmt_hrs(val) or None) if col in HRS_COLS else (
            None if (val is None or (isinstance(val, float) and np.isnan(val))) else val)

        dc = ws.cell(ri, j, cell_val)

        if col == "TIPO_CITA":
            if "SCAC" in tipo_v:
                dc.fill = TIPO_SCAC_FILL; dc.font = TIPO_SCAC_FONT
            else:
                dc.fill = TIPO_NUEV_FILL; dc.font = TIPO_NUEV_FONT
        elif col in ("CEDIS","NOMBRE_CEDIS") and row_fill:
            dc.fill = row_fill; dc.font = Font(bold=True, size=9)
        else:
            dc.font = Font(bold=(col in ("CITA","SW","RECIBO_HRS","LOS_HRS")), size=9)
            if row_fill: dc.fill = row_fill

        dc.alignment = Alignment(
            horizontal="left" if col in LEFT_COLS else "center",
            vertical="center")
        dc.border = BRD

# ─── Totales por CEDIS ────────────────────────────────────────────────────────
ri_sep = len(df) + 3
ws.row_dimensions[ri_sep].height = 8
ri_tot = ri_sep + 1

# Encabezado resumen
resumen_header = ws.cell(ri_tot, 1, "RESUMEN POR CEDIS")
resumen_header.fill = tot_fill; resumen_header.font = Font(bold=True, color="FFFFFF", size=9)
resumen_header.alignment = Alignment(horizontal="center", vertical="center")
resumen_header.border = BRD
ws.merge_cells(start_row=ri_tot, start_column=1, end_row=ri_tot, end_column=6)
ri_tot += 1

# Cabecera resumen
for j, lbl in enumerate(["CEDIS","NOMBRE","SCAC INCORRECTO","CITA NUEVA","TOTAL","SWs"], 1):
    c = ws.cell(ri_tot, j, lbl)
    c.fill = hdr_fill; c.font = Font(bold=True, color="FFFFFF", size=9)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BRD
ri_tot += 1

for cedis, nombre in CEDIS_GDL.items():
    sub   = df[df["CEDIS"] == cedis]
    n_sc  = int((sub["TIPO_CITA"] == "SCAC INCORRECTO").sum())
    n_cn  = int((sub["TIPO_CITA"] == "CITA NUEVA").sum())
    sws   = ", ".join(f"SW{int(s)}" for s in sorted(sub["SW"].dropna().unique()) if s)
    fill  = CEDIS_FILLS.get(cedis)
    for j, val in enumerate([cedis, nombre, n_sc or "", n_cn or "", len(sub) or "", sws or "—"], 1):
        dc = ws.cell(ri_tot, j, val)
        if fill: dc.fill = fill
        dc.font = Font(bold=(j <= 2), size=9)
        dc.alignment = Alignment(horizontal="center" if j != 2 else "left", vertical="center")
        dc.border = BRD
    ri_tot += 1

# Total general
for j, val in enumerate(["TOTAL", "", sum((df["TIPO_CITA"]=="SCAC INCORRECTO").values),
                          sum((df["TIPO_CITA"]=="CITA NUEVA").values), len(df), ""], 1):
    dc = ws.cell(ri_tot, j, val)
    dc.fill = tot_fill; dc.font = Font(bold=True, color="FFFFFF", size=9)
    dc.alignment = Alignment(horizontal="center", vertical="center")
    dc.border = BRD

# Anchos y freeze
for j, col in enumerate(COLS, 1):
    ws.column_dimensions[get_column_letter(j)].width = COL_W.get(col, 12)
ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 24
ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}2"

wb.save(OUT_FILE)
print(f"\nListo -> {OUT_FILE.name}")
